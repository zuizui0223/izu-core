#!/usr/bin/env python3
"""Acquire and inventory the source-native Wanshan–Yongxing Dryad workbook."""
from __future__ import annotations

import argparse
import hashlib
import http.cookiejar
import json
import re
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path
from typing import Any


USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36 "
    "izu-core-source-audit/1.0"
)


def make_opener() -> urllib.request.OpenerDirector:
    jar = http.cookiejar.CookieJar()
    return urllib.request.build_opener(
        urllib.request.HTTPCookieProcessor(jar),
        urllib.request.HTTPRedirectHandler(),
    )


def request_bytes(
    opener: urllib.request.OpenerDirector,
    url: str,
    *,
    accept: str = "application/octet-stream, */*;q=0.8",
    referer: str | None = None,
) -> bytes:
    headers = {
        "User-Agent": USER_AGENT,
        "Accept": accept,
        "Accept-Language": "en-US,en;q=0.8",
        "Cache-Control": "no-cache",
    }
    if referer:
        headers["Referer"] = referer
    request = urllib.request.Request(url, headers=headers)
    with opener.open(request, timeout=120) as response:
        return response.read()


def request_json(opener: urllib.request.OpenerDirector, url: str) -> dict[str, Any] | list[Any]:
    payload = request_bytes(opener, url, accept="application/json")
    decoded = json.loads(payload.decode("utf-8"))
    if not isinstance(decoded, (dict, list)):
        raise ValueError(f"JSON response from {url} must be an object or array")
    return decoded


def recursive_values(value: object, key_name: str) -> list[object]:
    found: list[object] = []
    if isinstance(value, dict):
        for key, child in value.items():
            if key == key_name:
                found.append(child)
            found.extend(recursive_values(child, key_name))
    elif isinstance(value, list):
        for child in value:
            found.extend(recursive_values(child, key_name))
    return found


def latest_version_id(versions: object) -> int | None:
    if isinstance(versions, dict):
        embedded = versions.get("_embedded")
        if isinstance(embedded, dict):
            rows = embedded.get("stash:versions")
            if isinstance(rows, list):
                ids = [row.get("id") for row in rows if isinstance(row, dict)]
                integers = [value for value in ids if isinstance(value, int)]
                if integers:
                    return max(integers)
    candidates = recursive_values(versions, "id")
    integers = [value for value in candidates if isinstance(value, int)]
    return max(integers) if integers else None


def candidate_file_ids(file_listing: object, target_filename: str) -> list[int]:
    ids: list[int] = []
    embedded = file_listing.get("_embedded") if isinstance(file_listing, dict) else None
    rows: list[dict[str, Any]] = []
    if isinstance(embedded, dict):
        for value in embedded.values():
            if isinstance(value, list):
                rows.extend(item for item in value if isinstance(item, dict))
    for row in rows:
        path = str(row.get("path") or row.get("name") or "")
        if path == target_filename or path.lower().endswith(".xlsx"):
            file_id = row.get("id")
            if isinstance(file_id, int):
                ids.append(file_id)
    return ids


def extract_linkset_download_urls(linkset: object) -> list[str]:
    """Extract current public file-stream URLs from a Dryad linkset response."""
    urls: list[str] = []
    for href in recursive_values(linkset, "href"):
        if isinstance(href, str) and "/downloads/file_stream/" in href and href not in urls:
            urls.append(href)
    return urls


def zip_assembly_candidates(
    opener: urllib.request.OpenerDirector,
    *,
    version_id: int,
    target_filename: str,
    errors: list[dict[str, str]],
) -> tuple[list[str], object]:
    """Ask the public UI endpoint for permanent presigned file URLs."""
    payload: object = {}
    urls: list[str] = []
    endpoints = (
        f"https://datadryad.org/downloads/zip_assembly_info/{version_id}.json",
        f"https://datadryad.org/downloads/zip_assembly_info/{version_id}?format=json",
    )
    for endpoint in endpoints:
        try:
            payload = request_json(opener, endpoint)
            rows = payload if isinstance(payload, list) else []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                filename = str(row.get("filename") or "")
                url = row.get("url")
                if isinstance(url, str) and (
                    filename == target_filename or filename.lower().endswith(".xlsx")
                ) and url not in urls:
                    urls.append(url)
            if urls:
                break
        except Exception as error:
            errors.append({"stage": "zip_assembly_info", "url": endpoint, "error": repr(error)})
    return urls, payload


def is_xlsx(payload: bytes) -> bool:
    if len(payload) < 1000 or not payload.startswith(b"PK"):
        return False
    import io

    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            names = set(archive.namelist())
    except zipfile.BadZipFile:
        return False
    return "[Content_Types].xml" in names and any(name.startswith("xl/") for name in names)


def workbook_preview(path: Path) -> dict[str, object]:
    try:
        import openpyxl
    except ImportError:
        return {"preview_error": "openpyxl_not_installed"}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets: list[dict[str, object]] = []
    for worksheet in workbook.worksheets:
        preview = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True)):
            preview.append(list(row[:8]))
            if row_index >= 4:
                break
        sheets.append(
            {
                "sheet": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "preview_first_8_columns": preview,
            }
        )
    workbook.close()
    return {"sheets": sheets}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--config",
        type=Path,
        default=Path("config/wanshan_yongxing_dryad_source.json"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("artifacts/wanshan_yongxing_dryad"),
    )
    args = parser.parse_args()

    config = json.loads(args.config.read_text(encoding="utf-8"))
    args.output_dir.mkdir(parents=True, exist_ok=True)
    opener = make_opener()
    errors: list[dict[str, str]] = []

    try:
        request_bytes(
            opener,
            str(config["landing_page_url"]),
            accept="text/html,application/xhtml+xml",
        )
    except Exception as error:
        errors.append({"stage": "landing_page_warmup", "error": repr(error)})

    metadata: object = {}
    versions: object = {}
    file_listing: object = {}
    linkset: object = {}
    zip_info: object = {}
    try:
        metadata = request_json(opener, str(config["api_dataset_url"]))
    except Exception as error:
        errors.append({"stage": "dataset_metadata", "error": repr(error)})
    try:
        versions = request_json(opener, str(config["api_versions_url"]))
    except Exception as error:
        errors.append({"stage": "dataset_versions", "error": repr(error)})

    download_candidates = list(map(str, config.get("public_download_candidates", [])))

    linkset_urls = (
        str(config.get("linkset_json_url") or ""),
        str(config["landing_page_url"]).rstrip("/") + "/linkset.json",
    )
    for linkset_url in dict.fromkeys(url for url in linkset_urls if url):
        try:
            linkset = request_json(opener, linkset_url)
            for url in extract_linkset_download_urls(linkset):
                if url not in download_candidates:
                    download_candidates.append(url)
            break
        except Exception as error:
            errors.append({"stage": "linkset", "url": linkset_url, "error": repr(error)})

    file_ids: list[int] = []
    version_id = latest_version_id(versions)
    if version_id is not None:
        files_url = f"https://datadryad.org/api/v2/versions/{version_id}/files"
        try:
            file_listing = request_json(opener, files_url)
            file_ids.extend(
                candidate_file_ids(file_listing, str(config["known_file"]["filename"]))
            )
        except Exception as error:
            errors.append({"stage": "version_files", "url": files_url, "error": repr(error)})

        presigned_urls, zip_info = zip_assembly_candidates(
            opener,
            version_id=version_id,
            target_filename=str(config["known_file"]["filename"]),
            errors=errors,
        )
        for url in presigned_urls:
            if url not in download_candidates:
                download_candidates.insert(0, url)

    known_id = int(config["known_file"]["file_id"])
    if known_id not in file_ids:
        file_ids.append(known_id)

    for file_id in file_ids:
        for url in (
            f"https://datadryad.org/downloads/file_stream/{file_id}",
            f"https://datadryad.org/stash/downloads/file_stream/{file_id}",
            f"https://datadryad.org/api/v2/files/{file_id}/download",
        ):
            if url not in download_candidates:
                download_candidates.append(url)

    payload: bytes | None = None
    successful_url: str | None = None
    for url in download_candidates:
        try:
            candidate = request_bytes(
                opener,
                url,
                accept=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                    "application/zip,application/octet-stream;q=0.9,*/*;q=0.8"
                ),
                referer=str(config["landing_page_url"]),
            )
            if not is_xlsx(candidate):
                sample = re.sub(r"\s+", " ", candidate[:160].decode("utf-8", errors="replace"))
                raise ValueError(f"response is not an xlsx workbook; prefix={sample!r}")
            payload = candidate
            successful_url = url
            break
        except Exception as error:
            errors.append({"stage": "download", "url": url, "error": repr(error)})

    if payload is None or successful_url is None:
        diagnostic = {
            "source_id": config["source_id"],
            "dataset_doi": config["dataset_doi"],
            "version_id": version_id,
            "api_dataset_metadata": metadata,
            "api_versions_metadata": versions,
            "api_file_listing": file_listing,
            "linkset_metadata": linkset,
            "zip_assembly_info": zip_info,
            "download_candidates": download_candidates,
            "errors": errors,
        }
        diagnostic_path = args.output_dir / "acquisition_errors.json"
        diagnostic_path.write_text(
            json.dumps(diagnostic, indent=2, ensure_ascii=False, default=str) + "\n",
            encoding="utf-8",
        )
        print(json.dumps(diagnostic, indent=2, ensure_ascii=False, default=str))
        raise RuntimeError(f"unable to acquire Dryad workbook; see {diagnostic_path}")

    destination = args.output_dir / str(config["known_file"]["filename"])
    destination.write_bytes(payload)
    preview = workbook_preview(destination)
    summary = {
        "source_id": config["source_id"],
        "article_doi": config["article_doi"],
        "dataset_doi": config["dataset_doi"],
        "successful_download_url": successful_url,
        "filename": destination.name,
        "size_downloaded": len(payload),
        "sha256": hashlib.sha256(payload).hexdigest(),
        "api_dataset_metadata": metadata,
        "api_versions_metadata": versions,
        "api_file_listing": file_listing,
        "linkset_metadata": linkset,
        "zip_assembly_info": zip_info,
        "download_attempt_errors": errors,
        **preview,
        "claim_boundary": config["claim_boundary"],
    }
    (args.output_dir / "source_inventory.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8"
    )
    print(f"downloaded: {destination} ({len(payload)} bytes)")
    print(f"sha256: {summary['sha256']}")
    print(f"sheets: {len(summary.get('sheets', []))}")


if __name__ == "__main__":
    main()
