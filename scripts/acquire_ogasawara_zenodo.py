#!/usr/bin/env python3
"""Acquire and inventory the Ogasawara pollination-network Zenodo record."""
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import re
import urllib.error
import urllib.request
import zipfile
from pathlib import Path
from typing import Any


USER_AGENT = "izu-core-source-audit/1.0 (+https://github.com/zuizui0223/izu-core)"


def request_bytes(url: str, *, accept: str = "application/octet-stream, */*;q=0.8") -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT, "Accept": accept})
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            return response.read()
    except urllib.error.HTTPError as error:
        try:
            body = error.read(400).decode("utf-8", errors="replace")
        except Exception:
            body = ""
        raise RuntimeError(f"HTTP {error.code} {error.reason}: {body!r}") from error


def request_json(url: str) -> Any:
    return json.loads(request_bytes(url, accept="application/json").decode("utf-8"))


def safe_name(value: str) -> str:
    name = Path(str(value)).name
    name = re.sub(r"[^A-Za-z0-9._+()-]+", "_", name).strip("._")
    return name or "zenodo_file"


def checksum(payload: bytes, algorithm: str) -> str:
    digest = hashlib.new(algorithm)
    digest.update(payload)
    return digest.hexdigest()


def verify_zenodo_checksum(payload: bytes, declared: str | None) -> dict[str, object]:
    if not declared or ":" not in declared:
        return {"declared_checksum": declared, "verified": None}
    algorithm, expected = declared.split(":", 1)
    algorithm = algorithm.lower().strip()
    if algorithm not in hashlib.algorithms_available:
        return {"declared_checksum": declared, "verified": None, "reason": "unsupported_algorithm"}
    observed = checksum(payload, algorithm)
    return {
        "declared_checksum": declared,
        "observed_checksum": f"{algorithm}:{observed}",
        "verified": observed.lower() == expected.lower(),
    }


def preview_delimited(path: Path) -> dict[str, object]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = text[:8192]
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;").delimiter
    except csv.Error:
        pass
    rows = []
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    for index, row in enumerate(reader):
        rows.append(row[:30])
        if index >= 5:
            break
    return {"kind": "delimited", "delimiter": delimiter, "preview": rows}


def preview_xlsx(path: Path) -> dict[str, object]:
    try:
        import openpyxl
    except ImportError:
        return {"kind": "xlsx", "preview_error": "openpyxl_not_installed"}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for worksheet in workbook.worksheets:
        rows = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True)):
            rows.append([value for value in row[:30]])
            if index >= 5:
                break
        sheets.append(
            {
                "sheet": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "preview": rows,
            }
        )
    workbook.close()
    return {"kind": "xlsx", "sheets": sheets}


def safe_extract_zip(path: Path, destination: Path) -> list[str]:
    destination.mkdir(parents=True, exist_ok=True)
    members = []
    with zipfile.ZipFile(path) as archive:
        for member in archive.infolist():
            member_path = Path(member.filename)
            if member_path.is_absolute() or ".." in member_path.parts:
                continue
            members.append(member.filename)
            if member.is_dir():
                continue
            target = destination / member_path
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(archive.read(member))
    return members


def preview_file(path: Path, extract_root: Path) -> dict[str, object]:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".tsv", ".txt"}:
        return preview_delimited(path)
    if suffix in {".xlsx", ".xlsm"}:
        return preview_xlsx(path)
    if suffix == ".zip" or zipfile.is_zipfile(path):
        members = safe_extract_zip(path, extract_root / path.stem)
        return {"kind": "zip", "archive_members": members}
    return {"kind": "other"}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=Path("config/ogasawara_zenodo_source.json"))
    parser.add_argument("--output-dir", type=Path, default=Path("artifacts/ogasawara_zenodo"))
    args = parser.parse_args()

    config = json.loads(args.config.read_text(encoding="utf-8"))
    args.output_dir.mkdir(parents=True, exist_ok=True)
    raw_dir = args.output_dir / "files"
    extract_root = args.output_dir / "extracted"
    raw_dir.mkdir(parents=True, exist_ok=True)

    try:
        record = request_json(str(config["api_url"]))
    except Exception as error:
        failure = {"status": "metadata_acquisition_failed", "error": repr(error), **config}
        (args.output_dir / "acquisition_failure.json").write_text(
            json.dumps(failure, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
        )
        raise

    files = record.get("files") or []
    inventory = []
    errors = []
    for item in files:
        key = str(item.get("key") or item.get("filename") or f"file_{item.get('id')}")
        links = item.get("links") or {}
        content_url = str(links.get("content") or links.get("self") or "")
        row: dict[str, object] = {
            "id": item.get("id"),
            "key": key,
            "size_reported": item.get("size"),
            "checksum_reported": item.get("checksum"),
            "content_url": content_url,
        }
        if not content_url:
            row["status"] = "missing_content_url"
            inventory.append(row)
            continue
        try:
            payload = request_bytes(content_url)
            destination = raw_dir / safe_name(key)
            destination.write_bytes(payload)
            row.update(
                {
                    "status": "downloaded",
                    "local_name": destination.name,
                    "size_downloaded": len(payload),
                    "sha256": hashlib.sha256(payload).hexdigest(),
                    "checksum_verification": verify_zenodo_checksum(payload, item.get("checksum")),
                    "preview": preview_file(destination, extract_root),
                }
            )
        except Exception as error:
            row["status"] = "download_failed"
            row["error"] = repr(error)
            errors.append({"key": key, "error": repr(error)})
        inventory.append(row)

    metadata = record.get("metadata") or {}
    summary = {
        "status": "acquired" if any(row.get("status") == "downloaded" for row in inventory) else "no_files_acquired",
        "source_id": config["source_id"],
        "record_id": config["record_id"],
        "dataset_doi": config["dataset_doi"],
        "zenodo_record_id": record.get("id"),
        "title": metadata.get("title"),
        "publication_date": metadata.get("publication_date"),
        "version": metadata.get("version"),
        "license": metadata.get("license"),
        "resource_type": metadata.get("resource_type"),
        "n_files": len(inventory),
        "n_downloaded": sum(row.get("status") == "downloaded" for row in inventory),
        "files": inventory,
        "errors": errors,
        "expected_design_context": config["expected_design_context"],
        "claim_boundary": config["claim_boundary"],
    }
    (args.output_dir / "source_inventory.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8"
    )
    print(f"record: {summary['zenodo_record_id']}")
    print(f"files downloaded: {summary['n_downloaded']}/{summary['n_files']}")
    if summary["n_downloaded"] == 0:
        raise RuntimeError("Zenodo record exposed no downloadable files")


if __name__ == "__main__":
    main()
