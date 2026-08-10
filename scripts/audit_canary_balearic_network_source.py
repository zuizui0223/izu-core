#!/usr/bin/env python3
"""Audit Canary–Balearic supplementary tables without inventing network roles.

The source paper compares two oceanic-island communities in the Canary Islands
with two continental-island communities in the Balearic Islands.  This audit
only discovers source-native tables that could contain interaction matrices,
long interaction records, network metrics, or community metadata.  Admission to
analysis requires explicit community identity, geological-origin mapping,
interaction semantics, and sampling information.
"""
from __future__ import annotations

import argparse
import csv
import json
import math
import re
from pathlib import Path
from typing import Iterable, Sequence


ROLE_ALIASES = {
    "community": ("community", "network", "site", "locality", "location", "plot"),
    "island": ("island", "island_name"),
    "archipelago": ("archipelago", "island_group", "island_system"),
    "geological_origin": ("origin", "island_type", "geological", "oceanic", "continental"),
    "plant": ("plant", "plant_species", "floral_species", "flowering_species"),
    "pollinator": ("pollinator", "visitor", "flower_visitor", "animal_species"),
    "interaction_weight": ("visit", "frequency", "interaction", "count", "weight", "abundance"),
    "sampling_effort": ("effort", "observation_time", "hours", "transect", "sampling"),
    "connectance": ("connectance",),
    "nestedness": ("nestedness", "wnodf", "nodf"),
    "specialisation": ("specialisation", "specialization", "h2"),
    "interaction_diversity": ("interaction_diversity", "shannon", "diversity_of_interactions"),
    "species_richness": ("species_richness", "number_of_species", "plant_richness", "pollinator_richness"),
    "niche_overlap": ("niche_overlap", "overlap"),
}

NETWORK_METRIC_ROLES = {
    "connectance",
    "nestedness",
    "specialisation",
    "interaction_diversity",
    "species_richness",
    "niche_overlap",
}


def norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().casefold()).strip("_")


def role_matches(headers: Sequence[object]) -> dict[str, list[str]]:
    normalized = [(norm(value), str(value or "")) for value in headers if str(value or "").strip()]
    output: dict[str, list[str]] = {}
    for role, aliases in ROLE_ALIASES.items():
        output[role] = [
            original
            for key, original in normalized
            if any(key == alias or alias in key for alias in aliases)
        ]
    return output


def first_table_block(rows: Iterable[Sequence[object]]) -> tuple[list[object], list[list[object]]]:
    iterator = iter(rows)
    headers: list[object] = []
    for row in iterator:
        values = list(row)
        if any(str(value or "").strip() for value in values):
            headers = values
            break
    sample: list[list[object]] = []
    for row in iterator:
        values = list(row)
        if any(str(value or "").strip() for value in values):
            sample.append(values)
        if len(sample) >= 20:
            break
    return headers, sample


def _is_number(value: object) -> bool:
    if value in (None, ""):
        return False
    try:
        number = float(value)
    except (TypeError, ValueError):
        return False
    return math.isfinite(number)


def classify_table(
    path: Path,
    sheet: str | None,
    headers: list[object],
    sample: list[list[object]],
    n_rows: int,
    n_columns: int,
    delimiter: str | None,
) -> dict[str, object]:
    roles = role_matches(headers)
    matched = {role for role, values in roles.items() if values}
    long_candidate = {"plant", "pollinator", "interaction_weight"}.issubset(matched)
    metric_candidate = bool(matched & NETWORK_METRIC_ROLES) and bool(
        matched & {"community", "island", "archipelago", "geological_origin"}
    )
    metadata_candidate = bool(
        {"community", "island"}.issubset(matched)
        or {"community", "geological_origin"}.issubset(matched)
    )

    first_header = norm(headers[0]) if headers else ""
    numeric_cells = 0
    observed_cells = 0
    for row in sample:
        for value in row[1:]:
            if value not in (None, ""):
                observed_cells += 1
                numeric_cells += int(_is_number(value))
    numeric_fraction = numeric_cells / observed_cells if observed_cells else 0.0
    matrix_candidate = (
        n_columns >= 4
        and not long_candidate
        and numeric_fraction >= 0.7
        and (
            "plant" in first_header
            or "species" in first_header
            or first_header in {"", "taxon"}
        )
    )

    return {
        "file": str(path),
        "sheet": sheet,
        "n_rows_reported": n_rows,
        "n_columns_reported": n_columns,
        "delimiter": delimiter,
        "headers": [str(value or "") for value in headers],
        "role_matches": roles,
        "long_interaction_candidate": long_candidate,
        "wide_interaction_matrix_candidate": matrix_candidate,
        "network_metric_table_candidate": metric_candidate,
        "community_metadata_candidate": metadata_candidate,
        "sample_numeric_fraction_after_first_column": numeric_fraction,
        "sample_preview": sample[:5],
    }


def read_delimited(path: Path) -> list[dict[str, object]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    delimiter = "\t" if path.suffix.casefold() == ".tsv" else ","
    try:
        delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",\t;").delimiter
    except csv.Error:
        pass
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    headers, sample = first_table_block(rows)
    return [
        classify_table(
            path,
            None,
            headers,
            sample,
            len(rows),
            max((len(row) for row in rows), default=0),
            delimiter,
        )
    ]


def read_xlsx(path: Path) -> list[dict[str, object]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    output = []
    for worksheet in workbook.worksheets:
        headers, sample = first_table_block(worksheet.iter_rows(values_only=True))
        output.append(
            classify_table(
                path,
                worksheet.title,
                headers,
                sample,
                worksheet.max_row,
                worksheet.max_column,
                None,
            )
        )
    workbook.close()
    return output


def read_docx(path: Path) -> list[dict[str, object]]:
    from docx import Document

    document = Document(path)
    output = []
    for index, table in enumerate(document.tables, start=1):
        rows = [[cell.text for cell in row.cells] for row in table.rows]
        headers, sample = first_table_block(rows)
        output.append(
            classify_table(
                path,
                f"docx_table_{index}",
                headers,
                sample,
                len(rows),
                max((len(row) for row in rows), default=0),
                None,
            )
        )
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=Path("artifacts/canary_balearic_oup"),
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("artifacts/canary_balearic_oup/schema_audit.json"),
    )
    args = parser.parse_args()

    tables: list[dict[str, object]] = []
    errors: list[dict[str, str]] = []
    for path in sorted(args.input_dir.rglob("*")):
        if not path.is_file() or path.name in {
            "source_inventory.json",
            "schema_audit.json",
            "acquisition_failure.json",
        }:
            continue
        try:
            suffix = path.suffix.casefold()
            if suffix in {".csv", ".tsv", ".txt"}:
                tables.extend(read_delimited(path))
            elif suffix in {".xlsx", ".xlsm"}:
                tables.extend(read_xlsx(path))
            elif suffix == ".docx":
                tables.extend(read_docx(path))
        except Exception as error:
            errors.append({"file": str(path), "error": repr(error)})

    summary = {
        "status": "tabular_schema_audited" if tables else "no_parseable_tabular_supplement",
        "n_tabular_units": len(tables),
        "n_long_interaction_candidates": sum(bool(row["long_interaction_candidate"]) for row in tables),
        "n_wide_interaction_matrix_candidates": sum(bool(row["wide_interaction_matrix_candidate"]) for row in tables),
        "n_network_metric_table_candidates": sum(bool(row["network_metric_table_candidate"]) for row in tables),
        "n_community_metadata_candidates": sum(bool(row["community_metadata_candidate"]) for row in tables),
        "tables": tables,
        "errors": errors,
        "next_gate": (
            "Verify source-defined community labels, map two Canary and two Balearic communities, preserve sampling effort, and distinguish raw interaction matrices from derived network metrics before any geological-origin contrast is calculated."
        ),
        "claim_boundary": (
            "Header and numeric-pattern matching are schema discovery only. A table is not admitted as a network until plant, visitor, community, interaction-weight and sampling semantics are source-resolved. Network structure is not pollinator effectiveness, effective dependency or reproductive success."
        ),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(summary, indent=2, ensure_ascii=False, default=str) + "\n",
        encoding="utf-8",
    )
    print(f"tabular units: {summary['n_tabular_units']}")
    print(f"long candidates: {summary['n_long_interaction_candidates']}")
    print(f"wide candidates: {summary['n_wide_interaction_matrix_candidates']}")
    print(f"metric candidates: {summary['n_network_metric_table_candidates']}")


if __name__ == "__main__":
    main()
