from __future__ import annotations

import csv
import hashlib
import io
import json
import os
import re
import shutil
import subprocess
import tempfile
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DESIGN = ROOT / "data/design/abm_v9_thousand_island_lake_source_gate_v1.json"
OUT = ROOT / "data/results/thousand_island_lake_2022_source_audit.json"
RAW_DIR = ROOT / "data/external/thousand_island_lake_2022"
ZENODO_API = "https://zenodo.org/api/records/6519751"
USER_AGENT = "izu-core-til-source-audit/1.0"
TEXT_SUFFIXES = {".txt", ".csv", ".tsv", ".r", ".rmd", ".md", ".readme"}
R_WORKSPACE_SUFFIXES = {".rdata", ".rda", ".rds"}
SPREADSHEET_SUFFIXES = {".xlsx"}


def fetch_bytes(url: str) -> tuple[int | None, bytes | None, str | None]:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            return int(response.status), response.read(), None
    except urllib.error.HTTPError as exc:
        return int(exc.code), None, str(exc)
    except Exception as exc:
        return None, None, f"{type(exc).__name__}: {exc}"


def sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def md5(payload: bytes) -> str:
    return hashlib.md5(payload).hexdigest()


def decode_text(payload: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "gb18030", "latin-1"):
        try:
            return payload.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise RuntimeError("text source could not be decoded")


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").lower()).strip("_")


def role_candidates(headers: list[str]) -> dict[str, list[str]]:
    normalized = [normalize(value) for value in headers]
    rules = {
        "network_or_site": ("network", "web", "site", "island", "plot", "location"),
        "time": ("year", "date", "month", "season", "time", "round", "survey"),
        "plant": ("plant", "lower", "resource"),
        "pollinator": ("pollinator", "insect", "visitor", "higher", "bee"),
        "interaction_amount": ("visit", "interaction", "frequency", "freq", "count", "abundance", "weight"),
    }
    return {
        role: [header for header, key in zip(headers, normalized) if any(token in key for token in tokens)]
        for role, tokens in rules.items()
    }


PAIR_CODE = re.compile(r"^(?:po[^_]*_pl[^_]*|pl[^_]*_po[^_]*)$", re.IGNORECASE)
SPATIAL_NETWORK_CODE = re.compile(r"^(?:b\d+|s\d+)$", re.IGNORECASE)
TEMPORAL_NETWORK_CODE = re.compile(r"^t(?:19|20)\d{2}$", re.IGNORECASE)


def wide_pair_matrix_structure(headers: list[str], data_rows: list[list[str]]) -> dict:
    """Recognize pair-by-network count matrices without calculating outcomes."""
    if not headers or not data_rows:
        return {
            "raw_pair_wide_table_visible": False,
            "pair_identity_column": None,
            "spatial_network_columns": [],
            "temporal_network_columns": [],
            "quantitative_network_columns": [],
        }

    width = len(headers)
    nonempty_rows = [row for row in data_rows if any(str(value).strip() for value in row)]
    rectangular = bool(nonempty_rows) and all(len(row) == width for row in nonempty_rows)
    if not rectangular:
        return {
            "raw_pair_wide_table_visible": False,
            "pair_identity_column": None,
            "spatial_network_columns": [],
            "temporal_network_columns": [],
            "quantitative_network_columns": [],
        }
    inspected_rows = nonempty_rows

    pair_column = None
    pair_column_index = None
    for index, header in enumerate(headers):
        values = [str(row[index]).strip() for row in inspected_rows if str(row[index]).strip()]
        if not values:
            continue
        normalized_header = normalize(header)
        header_is_pair = normalized_header in {"int", "interaction", "pair", "link"}
        encoded_fraction = sum(bool(PAIR_CODE.fullmatch(value)) for value in values) / len(values)
        if encoded_fraction >= 0.8 and (header_is_pair or index == 0):
            pair_column = header or "<row_names>"
            pair_column_index = index
            break

    spatial_columns = [header for header in headers if SPATIAL_NETWORK_CODE.fullmatch(str(header).strip())]
    temporal_columns = [header for header in headers if TEMPORAL_NETWORK_CODE.fullmatch(str(header).strip())]
    candidate_network_columns = spatial_columns + temporal_columns
    quantitative_columns = []
    for header in candidate_network_columns:
        index = headers.index(header)
        values = [str(row[index]).strip() for row in inspected_rows]
        try:
            numeric = [float(value) for value in values]
        except ValueError:
            continue
        if numeric and all(value >= 0.0 for value in numeric):
            quantitative_columns.append(header)

    visible = (
        pair_column_index is not None
        and len(candidate_network_columns) >= 2
        and set(quantitative_columns) == set(candidate_network_columns)
    )
    return {
        "raw_pair_wide_table_visible": visible,
        "pair_identity_column": pair_column,
        "pair_identity_encoding": "pollinator_plant_code" if pair_column_index is not None else None,
        "rectangular_data_rows": rectangular,
        "inspected_data_rows": len(inspected_rows),
        "spatial_network_columns": spatial_columns,
        "temporal_network_columns": temporal_columns,
        "quantitative_network_columns": quantitative_columns,
    }


def sniff_delimited(path: Path) -> dict:
    payload = path.read_bytes()
    text, encoding = decode_text(payload)
    lines = text.splitlines()
    sample = "\n".join(lines[:40])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="\t,;")
        delimiter = dialect.delimiter
    except csv.Error:
        first = lines[0] if lines else ""
        if "\t" in first:
            delimiter = "\t"
        elif first.count(";") > first.count(","):
            delimiter = ";"
        else:
            delimiter = ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    best_row = None
    best_headers: list[str] = []
    best_score = -1
    for index, row in enumerate(rows[:12], start=1):
        headers = [str(value).strip() for value in row]
        roles = role_candidates(headers)
        score = sum(bool(values) for values in roles.values())
        if score > best_score:
            best_score = score
            best_row = index
            best_headers = headers
    roles = role_candidates(best_headers)
    wide_structure = wide_pair_matrix_structure(
        best_headers,
        rows[best_row:] if best_row is not None else [],
    )
    return {
        "format": "delimited_text",
        "encoding": encoding,
        "delimiter_repr": repr(delimiter),
        "row_count": len(rows),
        "column_count_first_row": len(rows[0]) if rows else 0,
        "candidate_header_row": best_row,
        "candidate_headers": best_headers,
        "field_role_candidates": roles,
        "raw_pair_long_table_visible": (
            bool(roles["plant"])
            and bool(roles["pollinator"])
            and bool(roles["interaction_amount"])
            and bool(roles["network_or_site"] or roles["time"])
        ),
        **wide_structure,
    }


def inspect_r_code(path: Path) -> dict:
    text, encoding = decode_text(path.read_bytes())
    lines = text.splitlines()
    tokens = (
        "read.csv", "read.table", "read.delim", "readxl", "load(", "readrds", "matrix", "network",
        "interaction", "plant", "pollinator", "year", "island", "site", "visit", "null_model",
    )
    structural = [
        {"line": index, "text": line[:500]}
        for index, line in enumerate(lines, start=1)
        if any(token in line.lower() for token in tokens)
    ][:250]
    return {
        "format": "r_source_code",
        "encoding": encoding,
        "line_count": len(lines),
        "structural_indicator_lines": structural,
    }


def inspect_xlsx(path: Path) -> dict:
    book = load_workbook(path, read_only=True, data_only=False)
    sheets = []
    for sheet in book.worksheets:
        preview = [list(row) for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 12), values_only=True)]
        best_row = None
        best_headers: list[str] = []
        best_score = -1
        for index, row in enumerate(preview, start=1):
            headers = [str(value).strip() if value is not None else "" for value in row]
            roles = role_candidates(headers)
            score = sum(bool(values) for values in roles.values())
            if score > best_score:
                best_score = score
                best_row = index
                best_headers = headers
        sheets.append({
            "name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "candidate_header_row": best_row,
            "candidate_headers": best_headers,
            "field_role_candidates": role_candidates(best_headers),
        })
    book.close()
    return {"format": "xlsx", "sheet_count": len(sheets), "sheets": sheets}


def list_rar_members(path: Path) -> dict:
    if shutil.which("lsar") is None:
        raise RuntimeError("lsar is required for RAR member inventory")
    proc = subprocess.run(
        ["lsar", "-json", str(path)],
        check=True,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    data = json.loads(proc.stdout)
    contents = data.get("lsarContents") or []
    members = []
    for row in contents:
        members.append({
            "name": row.get("XADFileName"),
            "size": row.get("XADFileSize"),
            "is_directory": bool(row.get("XADIsDirectory", False)),
        })
    return {"member_count": len(members), "members": members}


def extract_rar(path: Path, destination: Path) -> dict:
    if shutil.which("unar") is None:
        raise RuntimeError("unar is required for RAR extraction")
    destination.mkdir(parents=True, exist_ok=True)
    proc = subprocess.run(
        ["unar", "-q", "-f", "-o", str(destination), str(path)],
        check=False,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )
    extracted_files = [item for item in destination.rglob("*") if item.is_file()]
    if not extracted_files:
        detail = (proc.stderr or proc.stdout or "unar produced no files").strip()[-1000:]
        raise RuntimeError(f"unar extraction failed with exit {proc.returncode}: {detail}")
    return {
        "returncode": proc.returncode,
        "files_extracted": len(extracted_files),
        "nonzero_return_with_recovered_files": proc.returncode != 0,
        "stderr_tail": (proc.stderr or "").strip()[-1000:],
    }


def r_workspace_structure(path: Path) -> dict:
    if shutil.which("Rscript") is None:
        return {"format": "r_workspace", "inspection_error": "Rscript not installed"}
    script = r'''
args <- commandArgs(trailingOnly=TRUE)
f <- args[[1]]
cat("FILE\t", basename(f), "\n", sep="")
print_obj <- function(name, obj, depth=0) {
  cls <- paste(class(obj), collapse="|")
  d <- dim(obj)
  dimtxt <- if (is.null(d)) "" else paste(d, collapse="x")
  nms <- names(obj)
  nmstxt <- if (is.null(nms)) "" else paste(head(nms, 80), collapse="|")
  rn <- rownames(obj); cn <- colnames(obj)
  rntxt <- if (is.null(rn)) "" else paste(head(rn, 20), collapse="|")
  cntxt <- if (is.null(cn)) "" else paste(head(cn, 20), collapse="|")
  cat("OBJECT\t", name, "\t", cls, "\t", length(obj), "\t", dimtxt, "\t", nmstxt, "\t", rntxt, "\t", cntxt, "\n", sep="")
  if (depth < 2 && is.list(obj) && !is.data.frame(obj)) {
    lim <- min(length(obj), 160)
    if (lim > 0) {
      for (i in seq_len(lim)) {
        childname <- if (!is.null(names(obj)) && nzchar(names(obj)[i])) names(obj)[i] else as.character(i)
        child <- obj[[i]]
        childcls <- paste(class(child), collapse="|")
        childdim <- dim(child)
        childdimtxt <- if (is.null(childdim)) "" else paste(childdim, collapse="x")
        childrn <- rownames(child); childcn <- colnames(child)
        childrntxt <- if (is.null(childrn)) "" else paste(head(childrn, 10), collapse="|")
        childcntxt <- if (is.null(childcn)) "" else paste(head(childcn, 10), collapse="|")
        cat("ELEMENT\t", name, "\t", i, "\t", childname, "\t", childcls, "\t", length(child), "\t", childdimtxt, "\t", childrntxt, "\t", childcntxt, "\n", sep="")
      }
    }
  }
}
if (grepl("\\.rds$", tolower(f))) {
  obj <- readRDS(f)
  print_obj("RDS", obj, 0)
} else {
  e <- new.env(parent=emptyenv())
  loaded <- load(f, envir=e)
  cat("LOADED_NAMES\t", paste(loaded, collapse="|"), "\n", sep="")
  for (nm in loaded) print_obj(nm, get(nm, envir=e), 0)
}
'''
    with tempfile.NamedTemporaryFile("w", suffix=".R", delete=False, encoding="utf-8") as handle:
        handle.write(script)
        r_script = Path(handle.name)
    try:
        proc = subprocess.run(
            ["Rscript", str(r_script), str(path)],
            check=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=120,
        )
        lines = [line for line in proc.stdout.splitlines() if line.strip()]
        matrix_like = 0
        list_element_matrix_like = 0
        named_site_time_tokens = 0
        for line in lines:
            lower = line.lower()
            if "matrix" in lower or "data.frame" in lower:
                matrix_like += 1
            if line.startswith("ELEMENT\t") and ("matrix" in lower or "data.frame" in lower):
                list_element_matrix_like += 1
            if any(token in lower for token in ("year", "island", "site", "network", "2017", "2018", "2019")):
                named_site_time_tokens += 1
        return {
            "format": "r_workspace",
            "structure_lines": lines[:500],
            "matrix_or_dataframe_structure_line_count": matrix_like,
            "list_element_matrix_like_count": list_element_matrix_like,
            "site_time_name_token_line_count": named_site_time_tokens,
        }
    finally:
        try:
            os.unlink(r_script)
        except OSError:
            pass


def inspect_extracted_file(path: Path) -> dict:
    record = {
        "relative_path": path.name,
        "bytes": path.stat().st_size,
        "sha256": sha256(path.read_bytes()),
        "suffix": path.suffix.lower(),
    }
    suffix = path.suffix.lower()
    try:
        if suffix in {".csv", ".tsv", ".txt"}:
            record["inventory"] = sniff_delimited(path)
        elif suffix in {".r", ".rmd"}:
            record["inventory"] = inspect_r_code(path)
        elif suffix in SPREADSHEET_SUFFIXES:
            record["inventory"] = inspect_xlsx(path)
        elif suffix in R_WORKSPACE_SUFFIXES:
            record["inventory"] = r_workspace_structure(path)
        else:
            record["inventory"] = {"format": "opaque_or_unhandled", "suffix": suffix}
    except Exception as exc:
        record["inventory_error"] = f"{type(exc).__name__}: {exc}"
    return record


def inventory_signals(extracted: list[dict]) -> dict:
    long_tables = []
    wide_tables = []
    r_matrix_structures = []
    temporal_tokens = []
    spatial_tokens = []
    for row in extracted:
        inv = row.get("inventory") or {}
        if inv.get("raw_pair_long_table_visible"):
            long_tables.append(row["relative_path"])
        if inv.get("raw_pair_wide_table_visible"):
            wide_tables.append({
                "path": row["relative_path"],
                "pair_identity_column": inv.get("pair_identity_column"),
                "spatial_network_columns": inv.get("spatial_network_columns", []),
                "temporal_network_columns": inv.get("temporal_network_columns", []),
                "quantitative_network_columns": inv.get("quantitative_network_columns", []),
            })
        if inv.get("format") == "r_workspace" and (
            inv.get("list_element_matrix_like_count", 0) > 0
            or inv.get("matrix_or_dataframe_structure_line_count", 0) >= 2
        ):
            r_matrix_structures.append({
                "path": row["relative_path"],
                "matrix_lines": inv.get("matrix_or_dataframe_structure_line_count"),
                "matrix_elements": inv.get("list_element_matrix_like_count"),
                "site_time_tokens": inv.get("site_time_name_token_line_count"),
            })
        signal_content = {
            "relative_path": row.get("relative_path"),
            "candidate_headers": inv.get("candidate_headers", []),
            "structural_indicator_lines": inv.get("structural_indicator_lines", []),
            "structure_lines": inv.get("structure_lines", []),
        }
        text_blob = json.dumps(signal_content, ensure_ascii=False).lower()
        if any(token in text_blob for token in ("year", "2017", "2018", "2019", "temporal")):
            temporal_tokens.append(row["relative_path"])
        if any(token in text_blob for token in ("island", "site", "spatial", "network")):
            spatial_tokens.append(row["relative_path"])
    raw_pair_structure_visible = bool(long_tables or wide_tables or r_matrix_structures)
    repeated_network_identifier_signal = bool(
        spatial_tokens
        or any(row["spatial_network_columns"] for row in wide_tables)
        or any(row["temporal_network_columns"] for row in wide_tables)
    )
    temporal_identifier_signal = bool(
        temporal_tokens or any(row["temporal_network_columns"] for row in wide_tables)
    )
    return {
        "raw_pair_long_table_files": sorted(set(long_tables)),
        "raw_pair_wide_table_files": wide_tables,
        "r_workspace_matrix_structure_files": r_matrix_structures,
        "spatial_or_network_identifier_signal_files": sorted(set(spatial_tokens)),
        "temporal_identifier_signal_files": sorted(set(temporal_tokens)),
        "raw_pair_structure_visible": raw_pair_structure_visible,
        "repeated_network_identifier_signal_visible": repeated_network_identifier_signal,
        "temporal_identifier_signal_visible": temporal_identifier_signal,
    }


def write(payload: dict) -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
    print(json.dumps(payload, indent=2, ensure_ascii=False))


def main() -> None:
    design = json.loads(DESIGN.read_text())
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    status, metadata_bytes, error = fetch_bytes(ZENODO_API)
    if status != 200 or metadata_bytes is None:
        write({
            "schema_version": "1.0",
            "analysis": "thousand_island_lake_2022_source_audit",
            "status": "blocked_til_zenodo_metadata_not_recovered",
            "metadata_http_status": status,
            "metadata_error": error,
            "source_admission_succeeds": False,
            "target_metrics_calculated": False,
        })
        return
    metadata = json.loads(metadata_bytes)
    files = metadata.get("files") or []
    file_records = []
    extracted_records = []
    blocked = []

    with tempfile.TemporaryDirectory(prefix="til_source_gate_") as temporary:
        temp_root = Path(temporary)
        for file_row in files:
            key = str(file_row.get("key") or "").strip()
            links = file_row.get("links") or {}
            url = links.get("content") or links.get("self")
            checksum = str(file_row.get("checksum") or "")
            expected_md5 = checksum.split(":", 1)[1] if checksum.startswith("md5:") else None
            record = {
                "key": key,
                "url": url,
                "metadata_size": file_row.get("size"),
                "metadata_checksum": checksum,
            }
            if not key or not url:
                record["error"] = "missing key or content URL"
                blocked.append(key or "<unnamed>")
                file_records.append(record)
                continue
            file_status, payload, file_error = fetch_bytes(url)
            record["http_status"] = file_status
            record["error"] = file_error
            if file_status != 200 or payload is None:
                blocked.append(key)
                file_records.append(record)
                continue
            record.update({
                "bytes": len(payload),
                "md5": md5(payload),
                "sha256": sha256(payload),
                "size_match": file_row.get("size") is None or int(file_row["size"]) == len(payload),
                "md5_match": expected_md5 is None or expected_md5 == md5(payload),
            })
            archive_path = RAW_DIR / re.sub(r"[^A-Za-z0-9._-]+", "_", key)
            archive_path.write_bytes(payload)
            if archive_path.suffix.lower() == ".rar":
                try:
                    record["archive_inventory"] = list_rar_members(archive_path)
                    extract_dir = temp_root / archive_path.stem
                    record["archive_extraction"] = extract_rar(archive_path, extract_dir)
                    member_records = []
                    for extracted_path in sorted(path for path in extract_dir.rglob("*") if path.is_file()):
                        member = inspect_extracted_file(extracted_path)
                        member["relative_path"] = str(extracted_path.relative_to(extract_dir))
                        member_records.append(member)
                        extracted_records.append({"archive": key, **member})
                    record["extracted_member_inventory"] = member_records
                except Exception as exc:
                    record["archive_inventory_error"] = f"{type(exc).__name__}: {exc}"
            file_records.append(record)

    source_bytes_ok = bool(files) and not blocked and all(row.get("size_match") and row.get("md5_match") for row in file_records if row.get("bytes"))
    signals = inventory_signals(extracted_records)
    admission = (
        source_bytes_ok
        and signals["raw_pair_structure_visible"]
        and signals["repeated_network_identifier_signal_visible"]
    )
    status_name = (
        "source_admitted_til_raw_network_structure_before_v9_targets"
        if admission
        else "blocked_til_raw_pair_or_network_structure_incomplete"
    )
    write({
        "schema_version": "1.0",
        "analysis": "thousand_island_lake_2022_source_audit",
        "status": status_name,
        "zenodo_record": 6519751,
        "metadata_bytes": len(metadata_bytes),
        "metadata_sha256": sha256(metadata_bytes),
        "metadata_title": (metadata.get("metadata") or {}).get("title"),
        "file_count": len(files),
        "files": file_records,
        "blocked_files": blocked,
        "source_bytes_ok": source_bytes_ok,
        "extracted_file_count": len(extracted_records),
        "source_structure_signals": signals,
        "source_admission_succeeds": admission,
        "target_metrics_calculated": False,
        "source_provenance_boundary": design["source_provenance_boundary"],
        "temporal_boundary": design["source_only_gate"]["temporal_boundary"],
        "observation_boundary": design["source_only_gate"]["observation_boundary"],
        "claim_boundary": (
            "Source archive/member/object structure only. No network metric, support estimand, ecological outcome, empirical range, or ABM v9 predictive fit is calculated. "
            "A later reconstruction gate is still required even if raw/network structure is admitted here."
        ),
    })


if __name__ == "__main__":
    main()
