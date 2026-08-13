#!/usr/bin/env python3
"""Audit whether the Southwest Pacific animal response shape depends on data source.

This is an adversarial sensitivity, not an errors-in-variables correction.  The
source-native S1 workbook labels each colonisation pair as coming from online
databases or herbaria; S2 supplies the pair-level FI/FM means.  We ask whether
the below-isometry direct response shape persists after dropping all herbarium
pairs.  Source-method strata are not interpreted as randomized groups, and the
small herbarium stratum is not used to claim a method difference.
"""
from __future__ import annotations

import argparse
import json
import math
import random
from pathlib import Path
from statistics import mean

from openpyxl import load_workbook


def _rows(path: Path, sheet: str) -> list[dict[str, object]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet]
        iterator = worksheet.iter_rows(values_only=True)
        header = [str(value or "").strip() for value in next(iterator)]
        output = []
        for row in iterator:
            record = {header[index]: value for index, value in enumerate(row) if index < len(header)}
            output.append(record)
        return output
    finally:
        workbook.close()


def _numeric(value: object) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) and number > 0 else None


def ols_slope(rows: list[dict[str, object]]) -> float | None:
    points = [(float(row["x"]), float(row["y"])) for row in rows]
    if len(points) < 2:
        return None
    xbar = mean(point[0] for point in points)
    ybar = mean(point[1] for point in points)
    denominator = sum((x - xbar) ** 2 for x, _ in points)
    if denominator <= 1e-14:
        return None
    return sum((x - xbar) * (y - ybar) for x, y in points) / denominator


def island_cluster_interval(
    rows: list[dict[str, object]], *, repetitions: int, seed: int
) -> tuple[float, float] | None:
    by_island: dict[str, list[dict[str, object]]] = {}
    for row in rows:
        by_island.setdefault(str(row["island"]), []).append(row)
    islands = sorted(by_island)
    if len(islands) < 2:
        return None
    rng = random.Random(seed)
    estimates: list[float] = []
    attempts = 0
    max_attempts = repetitions * 4
    while len(estimates) < repetitions and attempts < max_attempts:
        attempts += 1
        sampled: list[dict[str, object]] = []
        for island in rng.choices(islands, k=len(islands)):
            sampled.extend(by_island[island])
        estimate = ols_slope(sampled)
        if estimate is not None and math.isfinite(estimate):
            estimates.append(estimate)
    if len(estimates) < repetitions:
        raise RuntimeError("could not obtain requested non-degenerate island bootstrap replicates")
    estimates.sort()
    low_index = int(0.025 * (len(estimates) - 1))
    high_index = int(0.975 * (len(estimates) - 1))
    return estimates[low_index], estimates[high_index]


def load_animal_pairs(s1: Path, s2: Path) -> list[dict[str, object]]:
    source_rows = _rows(s1, "Data")
    source_by_pair = {
        int(row["Pair number"]): str(row.get("Data source") or "").strip()
        for row in source_rows
        if row.get("Pair number") not in (None, "")
    }
    flower_rows = _rows(s2, "Flower dataframe")
    output: list[dict[str, object]] = []
    for row in flower_rows:
        if row.get("Pair number") in (None, ""):
            continue
        if row.get("Syndrome") != 1:
            continue
        island_value = _numeric(row.get("FI"))
        mainland_value = _numeric(row.get("FM"))
        if island_value is None or mainland_value is None:
            continue
        pair_number = int(row["Pair number"])
        output.append(
            {
                "pair_number": pair_number,
                "source_method": source_by_pair.get(pair_number, ""),
                "island": str(row.get("Island") or "").strip(),
                "x": math.log10(mainland_value),
                "y": math.log10(island_value),
            }
        )
    return output


def summarize(rows: list[dict[str, object]], *, repetitions: int, seed: int) -> dict[str, object]:
    strata = {
        "all_animal": rows,
        "online_only": [row for row in rows if row["source_method"] == "Online databases"],
        "herbaria_only": [row for row in rows if row["source_method"] == "Herbaria"],
    }
    summaries: dict[str, object] = {}
    for index, (name, subset) in enumerate(strata.items()):
        slope = ols_slope(subset)
        interval = island_cluster_interval(
            subset, repetitions=repetitions, seed=seed + 1009 * index
        )
        summaries[name] = {
            "n_pairs": len(subset),
            "n_islands": len({str(row["island"]) for row in subset}),
            "direct_response_shape_slope": slope,
            "island_cluster_95": list(interval) if interval else None,
            "point_below_isometry": bool(slope is not None and slope < 1.0),
            "cluster_interval_wholly_below_isometry": bool(interval and interval[1] < 1.0),
        }
    online = summaries["online_only"]
    return {
        "schema_version": "1.0",
        "analysis_role": "source_method_adversarial_sensitivity_not_measurement_error_correction",
        "response_shape": "log10(island flower size) ~ log10(mainland flower size)",
        "pollination_subset": "source-coded animal-pollinated pairs",
        "bootstrap": {
            "cluster": "Island",
            "repetitions": repetitions,
            "seed": seed,
            "degenerate_draw_policy": "discard_and_resample",
        },
        "strata": summaries,
        "adversarial_reading": (
            "passes_online_only_below_isometry"
            if online["cluster_interval_wholly_below_isometry"]
            else "does_not_pass_online_only_below_isometry"
        ),
        "formal_consequence": {
            "compression_like_signal_requires_herbarium_pairs": False
            if online["cluster_interval_wholly_below_isometry"]
            else None,
            "source_method_difference_identified": False,
            "empirical_predictor_reliability_identified": False,
            "EIV_gate_opened": False,
            "formal_cross_system_admission_opened": False,
        },
        "claim_boundary": (
            "Persistence in the online-only subset argues against the full response shape being created solely by herbarium-source pairs. "
            "The strata are observational and imbalanced, so this analysis does not identify a source-method effect, estimate predictor reliability, or correct measurement error."
        ),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--s1", type=Path, required=True)
    parser.add_argument("--s2", type=Path, required=True)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--bootstrap-repetitions", type=int, default=5000)
    parser.add_argument("--seed", type=int, default=20260812)
    args = parser.parse_args()
    report = summarize(
        load_animal_pairs(args.s1, args.s2),
        repetitions=args.bootstrap_repetitions,
        seed=args.seed,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
