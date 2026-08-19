from __future__ import annotations

import hashlib
import json
import re
import urllib.parse
import urllib.request
from io import BytesIO
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
CONFIG = ROOT / "data/design/seychelles_kaiser_bunbury2017_weighted_source.json"
RAW_DIR = ROOT / "data/external/seychelles_kaiser_bunbury2017"
OUT = ROOT / "data/results/seychelles_kaiser_bunbury2017_weighted_source_audit.json"
WAYBACK_AVAILABILITY = "https://archive.org/wayback/available"


def get_bytes(url: str, timeout: int = 90) -> bytes:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": "izu-core-source-audit/1.0",
            "Accept": "text/html,application/json,application/vnd.ms-excel,application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,*/*",
        },
    )
    # Keep normal certificate + hostname verification enabled for every live
    # source and archive request. Transport failures are recorded, never bypassed.
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()


def sha256(data: bytes) -> str:
    h = hashlib.sha256()
    h.update(data)
    return h.hexdigest()


def hrefs_from_html(page_url: str, html: bytes) -> list[str]:
    text = html.decode("latin-1", errors="replace")
    hrefs = re.findall(r'href\s*=\s*["\']([^"\']+)["\']', text, flags=re.I)
    out = []
    for href in hrefs:
        absolute = urllib.parse.urljoin(page_url, href)
        lower = urllib.parse.unquote(absolute).lower()
        if any(token in lower for token in (".xls", ".xlsx", "excel", "no.visits", "visitfreq")):
            out.append(absolute)
    return list(dict.fromkeys(out))


def wayback_raw_snapshot_for_exact_url(original_url: str) -> dict | None:
    query = urllib.parse.urlencode({"url": original_url})
    api_url = f"{WAYBACK_AVAILABILITY}?{query}"
    payload = json.loads(get_bytes(api_url, timeout=90).decode("utf-8"))
    closest = (payload.get("archived_snapshots") or {}).get("closest") or {}
    if not closest.get("available") or str(closest.get("status")) != "200":
        return None
    timestamp = str(closest.get("timestamp") or "")
    snapshot_url = str(closest.get("url") or "")
    if not timestamp or not snapshot_url:
        return None
    # id_ asks Wayback for the archived payload rather than its replay HTML.
    raw_url = "https://web.archive.org/web/" + timestamp + "id_/" + original_url
    return {
        "availability_api_url": api_url,
        "timestamp": timestamp,
        "closest_replay_url": snapshot_url,
        "raw_snapshot_url": raw_url,
    }


def inspect_excel(name: str, data: bytes) -> dict:
    lower = name.lower()
    if data[:2] == b"PK" or lower.endswith(".xlsx"):
        import openpyxl

        book = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
        sheets = []
        for ws in book.worksheets:
            preview = []
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row or 0, 12), values_only=True):
                preview.append([x for x in row[:14]])
            sheets.append({"name": ws.title, "nrows": ws.max_row, "ncols": ws.max_column, "preview": preview})
        book.close()
        return {"format": "xlsx", "sheets": sheets}

    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    sheets = []
    for ws in book.sheets():
        preview = []
        for i in range(min(ws.nrows, 12)):
            preview.append([ws.cell_value(i, j) for j in range(min(ws.ncols, 14))])
        sheets.append({"name": ws.name, "nrows": ws.nrows, "ncols": ws.ncols, "preview": preview})
    return {"format": "xls", "sheets": sheets}


def source_label(filename: str, inspection: dict) -> str:
    tokens = [filename.lower()]
    tokens.extend(str(sheet.get("name", "")).lower() for sheet in inspection.get("sheets", []))
    joined = " | ".join(tokens)
    has_visitfreq = "visitfreq" in joined or "visit freq" in joined or "visitation frequency" in joined
    has_no_visits = "no.visits" in joined or "no visits" in joined or "no_visits" in joined
    if has_visitfreq and has_no_visits:
        return "combined_workbook_with_visitfreq_and_no_visits_source_labels"
    if has_visitfreq:
        return "visitfreq_candidate_by_source_label"
    if has_no_visits:
        return "no_visits_candidate_by_source_label"
    return "unclassified_excel_candidate"


def safe_local_name(page_index: int, link_index: int, url: str, suffix: str) -> str:
    raw = Path(urllib.parse.unquote(urllib.parse.urlparse(url).path)).name
    raw = re.sub(r"[^A-Za-z0-9._-]+", "_", raw).strip("_")
    if not raw or not raw.lower().endswith((".xls", ".xlsx")):
        raw = (raw or "source") + suffix
    return f"p{page_index:02d}_l{link_index:02d}_{raw}"


def validate_excel_payload(data: bytes) -> str:
    if len(data) < 1024:
        raise RuntimeError(f"payload too small: {len(data)}")
    if data[:8] == bytes.fromhex("D0CF11E0A1B11AE1"):
        return ".xls"
    if data[:2] == b"PK":
        return ".xlsx"
    raise RuntimeError(f"not an Excel payload: magic={data[:8]!r}")


def recover_exact_link(url: str) -> tuple[bytes, dict]:
    live_error = None
    try:
        data = get_bytes(url)
        validate_excel_payload(data)
        return data, {"transport": "live_exact_href", "url": url}
    except Exception as exc:
        live_error = repr(exc)

    archive = wayback_raw_snapshot_for_exact_url(url)
    if archive is None:
        raise RuntimeError(f"live exact href failed ({live_error}); no Wayback 200 snapshot for exact URL")
    data = get_bytes(archive["raw_snapshot_url"], timeout=120)
    validate_excel_payload(data)
    return data, {
        "transport": "internet_archive_exact_url_snapshot",
        "original_url": url,
        "live_error": live_error,
        **archive,
    }


def main() -> None:
    config = json.loads(CONFIG.read_text())
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    source_pages = config.get("source_pages") or [
        {"url": config["iwdb_page"], "role": "legacy single source page", "verification": "legacy config"}
    ]
    payload = {
        "schema_version": "1.2",
        "analysis": "seychelles_kaiser_bunbury2017_weighted_source_audit",
        "article_doi": config["article_doi"],
        "primary_tierb_weight_family": config["primary_tierb_weight_family"],
        "secondary_sensitivity_weight_family": config["secondary_sensitivity_weight_family"],
        "source_page_policy": config.get("source_page_policy"),
        "archive_policy": "If and only if a verified source page exposes an exact Excel href whose live bytes are unavailable, query Internet Archive availability for that exact href and accept only a status-200 archived payload with valid Excel magic. Do not alter host/path/filename to search for a better fit.",
        "status": "not_recovered",
        "source_page_attempts": [],
        "file_attempts": [],
    }

    recovered = []
    successful_pages = 0
    seen_file_sha = set()
    for page_index, page_spec in enumerate(source_pages):
        page = page_spec["url"]
        page_record = {
            "url": page,
            "role": page_spec.get("role"),
            "verification": page_spec.get("verification"),
            "status": "pending",
        }
        try:
            html = get_bytes(page)
            successful_pages += 1
            page_record.update({"status": "success", "bytes": len(html), "sha256": sha256(html)})
            (RAW_DIR / f"source_page_{page_index:02d}.html").write_bytes(html)
            hrefs = hrefs_from_html(page, html)
            page_record["candidate_data_links"] = hrefs
            for link_index, url in enumerate(hrefs):
                attempt = {"source_page": page, "url": url, "status": "pending"}
                try:
                    data, provenance = recover_exact_link(url)
                    suffix = validate_excel_payload(data)
                    digest = sha256(data)
                    filename = safe_local_name(page_index, link_index, url, suffix)
                    inspection = inspect_excel(filename, data)
                    label = source_label(filename, inspection)
                    duplicate = digest in seen_file_sha
                    if not duplicate:
                        seen_file_sha.add(digest)
                        (RAW_DIR / filename).write_bytes(data)
                        recovered.append({
                            "source_page": page,
                            "source_href": url,
                            "filename": filename,
                            "bytes": len(data),
                            "sha256": digest,
                            "source_label_class": label,
                            "recovery_provenance": provenance,
                            **inspection,
                        })
                    attempt.update({
                        "status": "success",
                        "bytes": len(data),
                        "sha256": digest,
                        "duplicate_sha256": duplicate,
                        "source_label_class": label,
                        "recovery_provenance": provenance,
                    })
                except Exception as exc:
                    attempt.update({"status": "failed", "error": repr(exc)})
                payload["file_attempts"].append(attempt)
        except Exception as exc:
            page_record.update({"status": "failed", "error": repr(exc)})
        payload["source_page_attempts"].append(page_record)

    visitfreq = [
        row for row in recovered
        if row["source_label_class"] in (
            "visitfreq_candidate_by_source_label",
            "combined_workbook_with_visitfreq_and_no_visits_source_labels",
        )
    ]
    no_visits = [
        row for row in recovered
        if row["source_label_class"] in (
            "no_visits_candidate_by_source_label",
            "combined_workbook_with_visitfreq_and_no_visits_source_labels",
        )
    ]
    if successful_pages == 0:
        status = "all_verified_source_pages_transport_failed"
        next_gate = "Search another independently verified public mirror; do not infer workbook URLs or source values."
    elif not recovered:
        status = "verified_source_page_recovered_no_excel_payload_recovered"
        next_gate = "Preserve the primary visitfreq validation as blocked; do not substitute no_visits."
    elif not visitfreq:
        status = "excel_candidates_recovered_primary_visitfreq_label_not_recovered"
        next_gate = "Preserve the primary visitfreq validation as blocked unless a source-labeled standardized workbook is independently recovered."
    else:
        status = "primary_visitfreq_source_labeled_candidate_recovered"
        next_gate = "Freeze workbook/sheet parsing roles from source labels and schema only, verify the complete 8-site x 8-month structure, and only then admit the primary visitfreq matrices for metric calculation."

    payload.update({
        "status": status,
        "successful_verified_source_pages": successful_pages,
        "recovered_excel_file_count": len(recovered),
        "recovered_excel_files": recovered,
        "primary_visitfreq_source_labeled_candidate_count": len(visitfreq),
        "secondary_no_visits_source_labeled_candidate_count": len(no_visits),
        "primary_visitfreq_candidate_sha256": [row["sha256"] for row in visitfreq],
        "next_gate": next_gate,
        "claim_boundary": config["claim_boundary"],
    })
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
    print(json.dumps(payload, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
