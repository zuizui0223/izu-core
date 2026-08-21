from __future__ import annotations

import hashlib
import io
import json
import re
import urllib.error
import urllib.request
import zipfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DESIGN = ROOT / "data/design/abm_v9_martinique_2025_source_gate_v1.json"
OUT = ROOT / "data/results/martinique_2025_v9_source_audit.json"
RAW_DIR = ROOT / "data/external/martinique_2025_v9"
USER_AGENT = "izu-core-martinique-v9-source-gate/1.1"


def fetch(url: str) -> tuple[int | None, bytes | None, str | None]:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=120) as response:
            return int(response.status), response.read(), None
    except urllib.error.HTTPError as exc:
        return int(exc.code), None, str(exc)
    except Exception as exc:
        return None, None, f"{type(exc).__name__}: {exc}"


def sha256(payload: bytes) -> str:
    return hashlib.sha256(payload).hexdigest()


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").lower()).strip("_")


def role_candidates(headers: list[str]) -> dict[str, list[str]]:
    normalized = [normalize(value) for value in headers]
    rules = {
        "site": ("site", "garden", "location", "plot", "locality"),
        "time": ("month", "date", "period", "biperiod", "bi_period", "season", "year", "sampling"),
        "plant": ("plant", "vegetal"),
        "pollinator": ("pollinator", "insect", "visitor", "arthropod"),
        "interaction_amount": ("interaction", "visit", "occurrence", "count", "number", "n_ind", "frequency", "abundance"),
        "floral_offer": ("floral", "flower", "open_flower", "floral_unit", "bloom", "inflorescence"),
        "effort": ("effort", "duration", "minute", "minutes", "hour", "transect", "census", "observation_time", "sampling_time", "h_start", "h_end"),
    }
    return {
        role: [header for header, key in zip(headers, normalized) if any(token in key for token in tokens)]
        for role, tokens in rules.items()
    }


def inspect_workbook(payload: bytes, name: str) -> dict:
    book = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    sheets = []
    for sheet in book.worksheets:
        preview = [
            list(row)
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 35), values_only=True)
        ]
        best = None
        best_score = -1
        for row_number, row in enumerate(preview, start=1):
            headers = [str(value).strip() if value is not None else "" for value in row]
            roles = role_candidates(headers)
            score = sum(bool(values) for values in roles.values())
            if score > best_score:
                best_score = score
                best = (row_number, headers, roles)
        row_number, headers, roles = best if best is not None else (None, [], role_candidates([]))

        interaction_identity_visible = bool(
            roles["site"] and roles["time"] and roles["plant"] and roles["pollinator"]
        )
        if interaction_identity_visible:
            interaction_representation = "explicit_amount" if roles["interaction_amount"] else "event_rows"
        else:
            interaction_representation = None
        interaction_candidate = interaction_identity_visible
        independent_floral_candidate = bool(
            roles["site"] and roles["time"] and roles["plant"] and roles["floral_offer"]
            and not roles["pollinator"]
        )
        exposure_candidate = bool(roles["site"] and roles["time"] and roles["effort"])
        sheets.append({
            "sheet": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "candidate_header_row": row_number,
            "candidate_headers": headers,
            "roles": roles,
            "repeated_interaction_candidate": interaction_candidate,
            "interaction_representation": interaction_representation,
            "independent_floral_offer_candidate": independent_floral_candidate,
            "sampling_exposure_candidate": exposure_candidate,
        })
    book.close()
    return {
        "name": name,
        "format": "xlsx",
        "sheet_count": len(sheets),
        "sheets": sheets,
    }


def docx_text(payload: bytes) -> str:
    with zipfile.ZipFile(io.BytesIO(payload)) as archive:
        xml = archive.read("word/document.xml").decode("utf-8", errors="replace")
    text = re.sub(r"<w:tab[^>]*/>", "\t", xml)
    text = re.sub(r"</w:p>", "\n", text)
    text = re.sub(r"<[^>]+>", "", text)
    text = text.replace("&amp;", "&").replace("&lt;", "<").replace("&gt;", ">")
    return re.sub(r"[ \t]+", " ", text)


def inspect_docx(payload: bytes, name: str) -> dict:
    text = docx_text(payload)
    lower = text.lower()
    protocol_terms = {
        "monthly": "month" in lower,
        "ten_sites": ("10 site" in lower) or ("ten site" in lower),
        "150m_transect": ("150 m" in lower) or ("150m" in lower),
        "60min_insect_observation": ("60 min" in lower) or ("60min" in lower) or ("60 minutes" in lower),
        "floral_offer": ("floral offer" in lower) or ("floral unit" in lower) or ("open flower" in lower),
        "plant_insect_interaction": ("plant-insect" in lower) or ("plant insect" in lower),
    }
    snippets = []
    for line in text.splitlines():
        line_lower = line.lower()
        if any(token in line_lower for token in ("floral", "flower", "60 min", "transect", "interaction", "month", "site")):
            snippets.append(line[:1000])
    return {
        "name": name,
        "format": "docx",
        "character_count": len(text),
        "protocol_terms": protocol_terms,
        "structural_snippets": snippets[:120],
    }


def transport_summary(design_sources: list[dict], records: list[dict]) -> dict:
    required_by_name = {row["name"]: bool(row.get("required", False)) for row in design_sources}
    recovered = {row["name"] for row in records if row.get("http_status") == 200 and row.get("bytes") is not None}
    required_names = [row["name"] for row in design_sources if row.get("required", False)]
    optional_names = [row["name"] for row in design_sources if not row.get("required", False)]
    blocked_required = [name for name in required_names if name not in recovered]
    blocked_optional = [name for name in optional_names if name not in recovered]
    return {
        "required_names": required_names,
        "optional_names": optional_names,
        "blocked_required_files": blocked_required,
        "blocked_optional_files": blocked_optional,
        "required_source_bytes_ok": not blocked_required and bool(required_names),
        "all_source_bytes_ok": not blocked_required and not blocked_optional,
        "required_by_name": required_by_name,
    }


def main() -> None:
    design = json.loads(DESIGN.read_text())
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    records = []
    workbooks = []
    readme = None

    for source in design["author_deposited_files"]:
        name = source["name"]
        status, payload, error = fetch(source["url"])
        record = {
            "name": name,
            "url": source["url"],
            "required": bool(source.get("required", False)),
            "role": source.get("role"),
            "http_status": status,
            "error": error,
        }
        if status != 200 or payload is None:
            records.append(record)
            continue
        record.update({"bytes": len(payload), "sha256": sha256(payload)})
        path = RAW_DIR / name
        path.write_bytes(payload)
        try:
            if name.lower().endswith(".xlsx"):
                inspection = inspect_workbook(payload, name)
                workbooks.append(inspection)
            elif name.lower().endswith(".docx"):
                inspection = inspect_docx(payload, name)
                readme = inspection
            else:
                inspection = {"name": name, "format": "unhandled"}
            record["inspection"] = inspection
        except Exception as exc:
            record["inspection_error"] = f"{type(exc).__name__}: {exc}"
        records.append(record)

    transport = transport_summary(design["author_deposited_files"], records)
    interaction_candidates = [
        {"file": book["name"], **sheet}
        for book in workbooks for sheet in book["sheets"]
        if sheet["repeated_interaction_candidate"]
    ]
    floral_candidates = [
        {"file": book["name"], **sheet}
        for book in workbooks for sheet in book["sheets"]
        if sheet["independent_floral_offer_candidate"]
    ]
    exposure_candidates = [
        {"file": book["name"], **sheet}
        for book in workbooks for sheet in book["sheets"]
        if sheet["sampling_exposure_candidate"]
    ]
    readme_protocol = readme["protocol_terms"] if readme else {}
    protocol_exposure_visible = bool(
        readme_protocol.get("monthly")
        and readme_protocol.get("150m_transect")
        and readme_protocol.get("60min_insect_observation")
    )
    required_source_bytes_ok = transport["required_source_bytes_ok"]
    repeated_interactions_visible = bool(interaction_candidates)
    independent_floral_offer_visible = bool(floral_candidates)
    exposure_visible = bool(exposure_candidates) or protocol_exposure_visible
    admission = bool(
        required_source_bytes_ok
        and repeated_interactions_visible
        and independent_floral_offer_visible
        and exposure_visible
    )

    if not required_source_bytes_ok:
        status_name = "blocked_martinique_required_primary_bytes_not_recovered"
    elif not repeated_interactions_visible:
        status_name = "blocked_martinique_repeated_interaction_schema_not_visible"
    elif not independent_floral_offer_visible:
        status_name = "blocked_martinique_independent_floral_offer_not_visible"
    elif not exposure_visible:
        status_name = "blocked_martinique_sampling_exposure_not_visible"
    else:
        status_name = "source_admitted_martinique_repeated_interactions_floral_offer_and_exposure_before_v9_targets"

    output = {
        "schema_version": "1.1",
        "analysis": "martinique_2025_v9_source_audit",
        "status": status_name,
        "source_admission_succeeds": admission,
        "required_source_bytes_ok": required_source_bytes_ok,
        "all_source_bytes_ok": transport["all_source_bytes_ok"],
        "blocked_required_files": transport["blocked_required_files"],
        "blocked_optional_files": transport["blocked_optional_files"],
        "files": records,
        "repeated_interaction_candidates": interaction_candidates,
        "independent_floral_offer_candidates": floral_candidates,
        "sampling_exposure_candidates": exposure_candidates,
        "readme_protocol": readme_protocol,
        "repeated_interactions_visible": repeated_interactions_visible,
        "independent_floral_offer_visible": independent_floral_offer_visible,
        "sampling_exposure_visible": exposure_visible,
        "primary_biological_scope": design["primary_biological_scope"],
        "target_metrics_calculated": False,
        "network_matrices_built": False,
        "interaction_events_aggregated": False,
        "v9_predictive_fit_calculated": False,
        "initial_parser_correction_note": (
            "The first source-only run required a separate interaction-amount column and all optional metadata bytes. "
            "Before any target calculation, the gate was corrected to recognize source-native event rows with site/time/plant/insect identity and to require only the two core biological workbooks."
        ),
        "claim_boundary": (
            "Source/schema admission only. Event rows are not aggregated here. Even a GO does not freeze monthly/bi-period reconstruction, identity handling, "
            "floral-opportunity definition, or any empirical target. Those require a separate pre-target reconstruction gate."
        ),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(output, indent=2, ensure_ascii=False) + "\n")
    print(json.dumps({
        "status": status_name,
        "source_admission": admission,
        "required_source_bytes_ok": required_source_bytes_ok,
        "all_source_bytes_ok": transport["all_source_bytes_ok"],
        "blocked_required_files": transport["blocked_required_files"],
        "blocked_optional_files": transport["blocked_optional_files"],
        "interaction_candidate_count": len(interaction_candidates),
        "floral_candidate_count": len(floral_candidates),
        "exposure_candidate_count": len(exposure_candidates),
        "readme_protocol": readme_protocol,
        "interaction_candidates": interaction_candidates,
        "floral_candidates": floral_candidates,
        "exposure_candidates": exposure_candidates,
        "file_hashes": {row["name"]: row.get("sha256") for row in records},
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
