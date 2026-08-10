#!/usr/bin/env python3
"""Acquire one public Dryad file through the resource zip-assembly metadata route."""
from __future__ import annotations

import argparse
import hashlib
import http.cookiejar
import json
import re
import urllib.request
import zipfile
from pathlib import Path
from typing import Any


USER_AGENT = (
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0 Safari/537.36 izu-core-source-audit/1.0"
)


def opener() -> urllib.request.OpenerDirector:
    return urllib.request.build_opener(
        urllib.request.HTTPCookieProcessor(http.cookiejar.CookieJar()),
        urllib.request.HTTPRedirectHandler(),
    )


def get_bytes(client: urllib.request.OpenerDirector, url: str, *, accept: str) -> bytes:
    request = urllib.request.Request(
        url,
        headers={
            "User-Agent": USER_AGENT,
            "Accept": accept,
            "Accept-Language": "en-US,en;q=0.8",
            "Cache-Control": "no-cache",
            "Referer": "https://datadryad.org/",
        },
    )
    with client.open(request, timeout=120) as response:
        return response.read()


def get_json(client: urllib.request.OpenerDirector, url: str) -> object:
    return json.loads(get_bytes(client, url, accept="application/json").decode("utf-8"))


def is_xlsx(payload: bytes) -> bool:
    if len(payload) < 1000 or not payload.startswith(b"PK"):
        return False
    import io

    try:
        with zipfile.ZipFile(io.BytesIO(payload)) as archive:
            names = set(archive.namelist())
    except zipfile.BadZipFile:
        return False
    return "[Content_Types].xml" in names and any(name.startswith("xl/") for name in names)


def preview_workbook(path: Path) -> list[dict[str, object]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    output: list[dict[str, object]] = []
    for worksheet in workbook.worksheets:
        preview: list[list[object]] = []
        for row_number, row in enumerate(worksheet.iter_rows(values_only=True)):
            preview.append(list(row[:8]))
            if row_number >= 4:
                break
        output.append(
            {
                "sheet": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "preview_first_8_columns": preview,
            }
        )
    workbook.close()
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--lock",
        type=Path,
        default=Path("config/wanshan_yongxing_dryad_version_lock.json"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("artifacts/wanshan_yongxing_dryad"),
    )
    args = parser.parse_args()

    lock = json.loads(args.lock.read_text(encoding="utf-8"))
    version_id = int(lock["resource_version_id"])
    target = str(lock["target_filename"])
    args.output_dir.mkdir(parents=True, exist_ok=True)
    client = opener()
    attempts: list[dict[str, object]] = []
    metadata: object = {}

    endpoints = [
        f"https://datadryad.org/downloads/zip_assembly_info/{version_id}.json",
        f"https://datadryad.org/downloads/zip_assembly_info/{version_id}?format=json",
        f"https://datadryad.org/downloads/zip_assembly_info/{version_id}",
    ]
    candidate_urls: list[str] = []
    for endpoint in endpoints:
        try:
            current = get_json(client, endpoint)
            attempts.append({"stage": "zip_assembly_info", "url": endpoint, "status": "success"})
            metadata = current
            rows = current if isinstance(current, list) else []
            for row in rows:
                if not isinstance(row, dict):
                    continue
                filename = str(row.get("filename") or "")
                url = row.get("url")
                if isinstance(url, str) and (filename == target or filename.lower().endswith(".xlsx")):
                    candidate_urls.append(url)
            if candidate_urls:
                break
        except Exception as error:
            attempts.append(
                {"stage": "zip_assembly_info", "url": endpoint, "status": "failure", "error": repr(error)}
            )

    payload: bytes | None = None
    successful_url: str | None = None
    for url in dict.fromkeys(candidate_urls):
        try:
            candidate = get_bytes(
                client,
                url,
                accept=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,"
                    "application/zip,application/octet-stream;q=0.9,*/*;q=0.8"
                ),
            )
            if not is_xlsx(candidate):
                prefix = re.sub(r"\s+", " ", candidate[:160].decode("utf-8", errors="replace"))
                raise ValueError(f"presigned response is not xlsx; prefix={prefix!r}")
            payload = candidate
            successful_url = url
            attempts.append({"stage": "presigned_download", "url": url, "status": "success"})
            break
        except Exception as error:
            attempts.append(
                {"stage": "presigned_download", "url": url, "status": "failure", "error": repr(error)}
            )

    if payload is None:
        diagnostic = {
            "dataset_doi": lock["dataset_doi"],
            "resource_version_id": version_id,
            "target_filename": target,
            "zip_assembly_metadata": metadata,
            "candidate_urls": candidate_urls,
            "attempts": attempts,
        }
        path = args.output_dir / "presigned_acquisition_errors.json"
        path.write_text(json.dumps(diagnostic, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
        print(json.dumps(diagnostic, indent=2, ensure_ascii=False))
        raise RuntimeError(f"public presigned acquisition failed; see {path}")

    destination = args.output_dir / target
    destination.write_bytes(payload)
    inventory = {
        "dataset_doi": lock["dataset_doi"],
        "resource_version_id": version_id,
        "target_file_id": lock["target_file_id"],
        "filename": target,
        "successful_download_url": successful_url,
        "size_downloaded": len(payload),
        "sha256": hashlib.sha256(payload).hexdigest(),
        "license": lock["license"],
        "source_locator": lock["source_locator"],
        "zip_assembly_metadata": metadata,
        "attempts": attempts,
        "sheets": preview_workbook(destination),
        "claim_boundary": lock["claim_boundary"],
    }
    (args.output_dir / "source_inventory.json").write_text(
        json.dumps(inventory, indent=2, ensure_ascii=False, default=str) + "\n",
        encoding="utf-8",
    )
    print(f"downloaded {destination} ({len(payload)} bytes)")
    print(f"sha256 {inventory['sha256']}")


if __name__ == "__main__":
    main()
