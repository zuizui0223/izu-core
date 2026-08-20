from __future__ import annotations

import csv
import hashlib
import io
import json
import re
import urllib.request
import zipfile
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DESIGN = ROOT / "data/design/abm_v5_menorca_source_gate_v1.json"
RAW_DIR = ROOT / "data/external/menorca2023"
OUT = ROOT / "data/results/menorca2023_figshare_source_audit.json"


def get_bytes(url: str, timeout: int = 120) -> bytes:
    request = urllib.request.Request(
        url,
        headers={"User-Agent": "izu-core-source-audit/1.0", "Accept": "application/json,application/octet-stream,*/*"},
    )
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.read()


def sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def md5(data: bytes) -> str:
    return hashlib.md5(data).hexdigest()


def likely_numeric(value: str) -> bool:
    text = str(value).strip()
    if not text:
        return False
    try:
        float(text)
        return True
    except ValueError:
        return False


def header_roles(headers: list[str]) -> dict:
    normalized = [re.sub(r"[^a-z0-9]+", "_", str(h).lower()).strip("_") for h in headers]
    families = {
        "plant": ("plant", "flower", "flora"),
        "pollinator": ("pollinator", "visitor", "insect", "bee"),
        "site": ("site", "locality", "location", "local_name", "zone"),
        "habitat": ("habitat", "vegetation"),
        "time": ("month", "date", "season", "subseason", "time", "sampling"),
        # FVR is the source's quantitative interaction field. Do not treat an
        # identity column such as Visitor_sp as an interaction weight merely
        # because its label contains "visitor".
        "weight": ("fvr", "visitation_rate", "visit_rate", "interaction_frequency", "interaction_freq", "frequency", "freq", "count", "n_visits", "no_visits"),
    }
    hits = {}
    for family, tokens in families.items():
        hits[family] = [
            original for original, normed in zip(headers, normalized)
            if any(token == normed or token in normed for token in tokens)
        ]
    return hits


def inspect_delimited(name: str, data: bytes) -> dict:
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:65536]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in sample else ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    rows = [row for row in rows if any(str(cell).strip() for cell in row)]
    if not rows:
        return {"format": "delimited_text", "delimiter": delimiter, "n_rows": 0, "n_columns": 0, "headers": [], "header_role_hits": {}}
    width = max(len(row) for row in rows)
    headers = [str(cell).strip() for cell in rows[0]]
    samples = {}
    for col_index, header in enumerate(headers):
        values = []
        for row in rows[1:]:
            if col_index >= len(row):
                continue
            value = str(row[col_index]).strip()
            if value and not likely_numeric(value) and value not in values:
                values.append(value)
            if len(values) >= 5:
                break
        if values:
            samples[header or f"column_{col_index+1}"] = values
    return {
        "format": "delimited_text",
        "delimiter": delimiter,
        "n_rows": len(rows),
        "n_columns_max": width,
        "headers": headers,
        "header_role_hits": header_roles(headers),
        "non_numeric_identifier_examples": samples,
    }


def inspect_excel(name: str, data: bytes) -> dict:
    lower = name.lower()
    sheets = []
    if lower.endswith(".xlsx") or data[:2] == b"PK":
        import openpyxl

        book = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        for sheet in book.worksheets:
            headers = []
            for row in sheet.iter_rows(min_row=1, max_row=1, values_only=True):
                headers = ["" if value is None else str(value).strip() for value in row]
            sheets.append({
                "name": sheet.title,
                "n_rows": sheet.max_row,
                "n_columns": sheet.max_column,
                "headers": headers,
                "header_role_hits": header_roles(headers),
                "source_defined_local_network": sheet.title != "Metaweb",
            })
        book.close()
        return {"format": "xlsx", "sheets": sheets}

    import xlrd

    book = xlrd.open_workbook(file_contents=data)
    for sheet in book.sheets():
        headers = [str(sheet.cell_value(0, i)).strip() for i in range(sheet.ncols)] if sheet.nrows else []
        sheets.append({
            "name": sheet.name,
            "n_rows": sheet.nrows,
            "n_columns": sheet.ncols,
            "headers": headers,
            "header_role_hits": header_roles(headers),
            "source_defined_local_network": sheet.name != "Metaweb",
        })
    return {"format": "xls", "sheets": sheets}


def inspect_zip(data: bytes) -> dict:
    members = []
    with zipfile.ZipFile(io.BytesIO(data)) as archive:
        for info in archive.infolist():
            if info.is_dir():
                continue
            row = {"name": info.filename, "bytes": info.file_size}
            suffix = Path(info.filename).suffix.lower()
            if suffix in (".csv", ".tsv", ".txt") and info.file_size <= 50_000_000:
                try:
                    row["schema"] = inspect_delimited(info.filename, archive.read(info))
                except Exception as exc:
                    row["schema_error"] = repr(exc)
            members.append(row)
    return {"format": "zip", "members": members}


def inspect_file(name: str, data: bytes) -> dict:
    suffix = Path(name).suffix.lower()
    if suffix in (".csv", ".tsv", ".txt"):
        return inspect_delimited(name, data)
    if suffix in (".xlsx", ".xls"):
        return inspect_excel(name, data)
    if suffix == ".zip" or data[:4] == b"PK\x03\x04" and not name.lower().endswith(".xlsx"):
        return inspect_zip(data)
    return {"format": suffix.lstrip(".") or "binary", "schema_inspected": False}


def main() -> None:
    design = json.loads(DESIGN.read_text())
    article_id = int(design["candidate_system"]["figshare_article_id"])
    metadata_url = f"https://api.figshare.com/v2/articles/{article_id}"
    metadata_bytes = get_bytes(metadata_url)
    metadata = json.loads(metadata_bytes.decode("utf-8"))
    if int(metadata.get("id")) != article_id:
        raise RuntimeError("Figshare article ID mismatch")

    public_files = metadata.get("files") or []
    if not public_files:
        raise RuntimeError("Figshare metadata exposes no public files")
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    files = []
    for file_meta in public_files:
        name = str(file_meta["name"])
        download_url = str(file_meta["download_url"])
        data = get_bytes(download_url)
        local_md5 = md5(data)
        supplied = file_meta.get("supplied_md5")
        computed = file_meta.get("computed_md5")
        if supplied and str(supplied).lower() != local_md5:
            raise RuntimeError(f"Figshare supplied MD5 mismatch for {name}")
        if computed and str(computed).lower() != local_md5:
            raise RuntimeError(f"Figshare computed MD5 mismatch for {name}")
        if file_meta.get("size") is not None and int(file_meta["size"]) != len(data):
            raise RuntimeError(f"Figshare byte-size mismatch for {name}")
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", name)
        (RAW_DIR / safe_name).write_bytes(data)
        files.append({
            "figshare_file_id": file_meta.get("id"),
            "name": name,
            "download_url": download_url,
            "bytes": len(data),
            "sha256": sha256(data),
            "local_md5": local_md5,
            "supplied_md5": supplied,
            "computed_md5": computed,
            "schema_audit": inspect_file(name, data),
        })

    local_network_sheets = []
    for file_row in files:
        for sheet in file_row.get("schema_audit", {}).get("sheets", []):
            if not sheet.get("source_defined_local_network"):
                continue
            roles = sheet.get("header_role_hits", {})
            if roles.get("plant") and roles.get("pollinator") and roles.get("weight"):
                local_network_sheets.append(sheet["name"])

    admitted = len(local_network_sheets) >= int(design["admission_requirements"]["minimum_local_network_units"])
    payload = {
        "schema_version": "1.1",
        "analysis": "menorca2023_figshare_source_schema_audit",
        "status": "source_admitted_nine_quantitative_local_network_sheets" if admitted else "source_blocked_insufficient_quantitative_local_network_schema",
        "figshare_article": {
            "id": article_id,
            "title": metadata.get("title"),
            "doi": metadata.get("doi"),
            "version": metadata.get("version"),
            "published_date": metadata.get("published_date"),
            "metadata_bytes": len(metadata_bytes),
            "metadata_sha256": sha256(metadata_bytes),
        },
        "public_file_count": len(files),
        "files": files,
        "source_defined_quantitative_local_network_count": len(local_network_sheets),
        "source_defined_quantitative_local_network_sheets": local_network_sheets,
        "source_admission_succeeds": admitted,
        "admitted_weight_column": "FVR" if admitted else None,
        "local_context_unit": "source-defined workbook sheet; no 3x3 site/habitat regrouping is inferred at this gate",
        "target_metrics_calculated": False,
        "source_gate_only": True,
        "next_gate": design["next_gate"],
        "claim_boundary": design["claim_boundary"],
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
    print(json.dumps({
        "status": payload["status"],
        "article": payload["figshare_article"],
        "files": [
            {"name": row["name"], "bytes": row["bytes"], "sha256": row["sha256"], "schema_audit": row["schema_audit"]}
            for row in files
        ],
        "local_network_sheets": local_network_sheets,
        "source_admission_succeeds": admitted,
        "admitted_weight_column": payload["admitted_weight_column"],
        "target_metrics_calculated": False,
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
