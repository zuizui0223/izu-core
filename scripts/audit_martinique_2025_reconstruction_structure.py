from __future__ import annotations

import io
import json
import math
from collections import Counter, defaultdict
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook

import acquire_audit_martinique_2025_v9_source as source_gate

ROOT = Path(__file__).resolve().parents[1]
SOURCE_DESIGN = ROOT / "data/design/abm_v9_martinique_2025_source_gate_v1.json"
OUT = ROOT / "data/results/martinique_2025_reconstruction_structure.json"
INTERACTION_NAME = "Plant_insect_interactions_former_names.xlsx"
SAMPLING_NAME = "Sampling_data.xlsx"
PROTOCOL_MINUTES_PER_SITE_PERIOD = 60


def fetch_required_bytes() -> dict[str, bytes]:
    design = json.loads(SOURCE_DESIGN.read_text())
    recovered: dict[str, bytes] = {}
    for row in design["author_deposited_files"]:
        if row["name"] not in {INTERACTION_NAME, SAMPLING_NAME}:
            continue
        status, payload, error = source_gate.fetch(row["url"])
        if status != 200 or payload is None:
            raise RuntimeError(f"required Martinique source unavailable: {row['name']} {status} {error}")
        recovered[row["name"]] = payload
    if set(recovered) != {INTERACTION_NAME, SAMPLING_NAME}:
        raise RuntimeError("required Martinique workbooks were not both recovered")
    return recovered


def rows_from_sheet(payload: bytes, sheet_name: str) -> tuple[list[str], list[dict[str, object]]]:
    book = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    if sheet_name not in book.sheetnames:
        raise RuntimeError(f"missing sheet {sheet_name}")
    sheet = book[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    header_row = next(iterator)
    headers = [str(value).strip() if value is not None else "" for value in header_row]
    rows = []
    for raw in iterator:
        if not any(value not in (None, "") for value in raw):
            continue
        rows.append({header: value for header, value in zip(headers, raw) if header})
    book.close()
    return headers, rows


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def canonical_month(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m")
    if isinstance(value, date):
        return value.strftime("%Y-%m")
    text = clean(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt).strftime("%Y-%m")
        except ValueError:
            pass
    if len(text) >= 7 and text[4] in "-/" and text[:4].isdigit():
        return text[:7].replace("/", "-")
    return None


def canonical_date(value: object) -> str | None:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = clean(value)
    return text or None


def minutes_since_midnight(value: object) -> float | None:
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return value.hour * 60 + value.minute + value.second / 60
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        number = float(value)
        if 0 <= number < 1:
            return number * 24 * 60
    text = clean(value)
    if not text:
        return None
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            parsed = datetime.strptime(text, fmt)
            return parsed.hour * 60 + parsed.minute + parsed.second / 60
        except ValueError:
            pass
    return None


def numeric(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def identity_structure(rows: list[dict[str, object]], best: str, genus: str, species: str, family: str) -> dict:
    blank_best = sum(not clean(row.get(best)) for row in rows)
    fallback_genus_species = sum(
        not clean(row.get(best)) and bool(clean(row.get(genus))) and bool(clean(row.get(species)))
        for row in rows
    )
    fallback_genus = sum(
        not clean(row.get(best)) and bool(clean(row.get(genus))) and not clean(row.get(species))
        for row in rows
    )
    fallback_family = sum(
        not clean(row.get(best)) and not clean(row.get(genus)) and bool(clean(row.get(family)))
        for row in rows
    )
    unresolved = sum(
        not clean(row.get(best)) and not clean(row.get(genus))
        and not clean(row.get(species)) and not clean(row.get(family))
        for row in rows
    )
    return {
        "row_count": len(rows),
        "nonblank_best_id_count": len({clean(row.get(best)) for row in rows if clean(row.get(best))}),
        "blank_best_id_rows": blank_best,
        "blank_best_with_genus_species_rows": fallback_genus_species,
        "blank_best_with_genus_only_rows": fallback_genus,
        "blank_best_with_family_only_rows": fallback_family,
        "fully_unresolved_rows": unresolved,
    }


def joint_interaction_identity_structure(rows: list[dict[str, object]]) -> dict:
    both_nonblank = []
    both_blank = []
    plant_only = []
    insect_only = []
    for row in rows:
        plant = clean(row.get("Plant_Best_ID"))
        insect = clean(row.get("Insect_Best_ID"))
        if plant and insect:
            both_nonblank.append(row)
        elif not plant and not insect:
            both_blank.append(row)
        elif plant:
            plant_only.append(row)
        else:
            insect_only.append(row)

    structural_fields = ("Period", "Date", "Site", "Transect", "H_start", "H_end", "Num_sp")
    blank_field_presence = {
        field: sum(bool(clean(row.get(field))) for row in both_blank)
        for field in structural_fields
    }
    num_sp_values = Counter(clean(row.get("Num_sp")) for row in both_blank if clean(row.get("Num_sp")))
    return {
        "both_best_ids_nonblank_rows": len(both_nonblank),
        "both_best_ids_blank_rows": len(both_blank),
        "plant_only_best_id_rows": len(plant_only),
        "insect_only_best_id_rows": len(insect_only),
        "both_blank_structural_field_nonblank_counts": blank_field_presence,
        "both_blank_num_sp_distinct_values": dict(sorted(num_sp_values.items())),
        "both_blank_source_context_count": len({
            (clean(row.get("Site")), clean(row.get("Period")))
            for row in both_blank if clean(row.get("Site")) and clean(row.get("Period"))
        }),
        "interpretation_boundary": (
            "This is a structural blank-pattern audit only. Blank-both rows are not yet assigned interaction weight or taxa."
        ),
    }


def context_keys(rows: list[dict[str, object]]) -> dict:
    sites = sorted({clean(row.get("Site")) for row in rows if clean(row.get("Site"))})
    periods = sorted({clean(row.get("Period")) for row in rows if clean(row.get("Period"))})
    dates = sorted({canonical_date(row.get("Date")) for row in rows if canonical_date(row.get("Date"))})
    months = sorted({canonical_month(row.get("Date")) for row in rows if canonical_month(row.get("Date"))})
    site_month = sorted({
        (clean(row.get("Site")), canonical_month(row.get("Date")))
        for row in rows if clean(row.get("Site")) and canonical_month(row.get("Date"))
    })
    site_period = sorted({
        (clean(row.get("Site")), clean(row.get("Period")))
        for row in rows if clean(row.get("Site")) and clean(row.get("Period"))
    })
    return {
        "sites": sites,
        "periods": periods,
        "dates": dates,
        "months": months,
        "site_month_keys": [list(key) for key in site_month],
        "site_period_keys": [list(key) for key in site_period],
        "site_count": len(sites),
        "period_count": len(periods),
        "date_count": len(dates),
        "month_count": len(months),
        "site_month_count": len(site_month),
        "site_period_count": len(site_period),
    }


def period_month_map(rows: list[dict[str, object]]) -> dict[str, list[str]]:
    mapping: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        period = clean(row.get("Period"))
        month = canonical_month(row.get("Date"))
        if period and month:
            mapping[period].add(month)
    return {key: sorted(values) for key, values in sorted(mapping.items())}


def site_month_coverage(rows: list[dict[str, object]]) -> dict[str, list[str]]:
    mapping: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        site = clean(row.get("Site"))
        month = canonical_month(row.get("Date"))
        if site and month:
            mapping[month].add(site)
    return {month: sorted(sites) for month, sites in sorted(mapping.items())}


def timing_field_structure(rows: list[dict[str, object]]) -> dict:
    start_present = 0
    end_present = 0
    both_present = 0
    parseable_both = 0
    positive_order = 0
    nonpositive_or_wrap = 0
    site_period_with_timing = set()
    for row in rows:
        start_raw = clean(row.get("H_start"))
        end_raw = clean(row.get("H_end"))
        start_present += bool(start_raw)
        end_present += bool(end_raw)
        if not start_raw or not end_raw:
            continue
        both_present += 1
        start = minutes_since_midnight(row.get("H_start"))
        end = minutes_since_midnight(row.get("H_end"))
        if start is None or end is None:
            continue
        parseable_both += 1
        if end > start:
            positive_order += 1
        else:
            nonpositive_or_wrap += 1
        site = clean(row.get("Site"))
        period = clean(row.get("Period"))
        if site and period:
            site_period_with_timing.add((site, period))
    return {
        "row_count": len(rows),
        "h_start_nonblank_rows": start_present,
        "h_end_nonblank_rows": end_present,
        "both_time_fields_nonblank_rows": both_present,
        "both_time_fields_parseable_rows": parseable_both,
        "event_rows_with_end_after_start": positive_order,
        "event_rows_nonpositive_or_wrap_order": nonpositive_or_wrap,
        "site_periods_with_parseable_event_timing": len(site_period_with_timing),
        "published_protocol_minutes_per_site_period": PROTOCOL_MINUTES_PER_SITE_PERIOD,
        "effort_rule_boundary": (
            "H_start/H_end are audited as event/provenance timing only and are never summed to estimate sampling effort. "
            "Prospective effort is the source-protocol fixed 60 min per Site×Period."
        ),
    }


def floral_structure(rows: list[dict[str, object]]) -> dict:
    values = [numeric(row.get("Nb_Floral_unit")) for row in rows]
    finite = [value for value in values if value is not None]
    missing_rows = [row for row, value in zip(rows, values) if value is None]
    quadrat_by_context: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in rows:
        site = clean(row.get("Site"))
        period = clean(row.get("Period"))
        quadrat = clean(row.get("Quadrat"))
        transect = clean(row.get("Transect"))
        if site and period and quadrat:
            quadrat_by_context[(site, period)].add(f"{transect}|{quadrat}")
    preview_fields = ("Period", "Date", "Site", "Transect", "Quadrat", "Plant_Best_ID", "Name_Floral_unit", "Nb_Floral_unit")
    missing_preview = [
        {field: clean(row.get(field)) for field in preview_fields}
        for row in missing_rows
    ]
    return {
        "nb_floral_unit_numeric_rows": len(finite),
        "nb_floral_unit_missing_or_nonnumeric_rows": len(missing_rows),
        "nb_floral_unit_negative_rows": sum(value < 0 for value in finite),
        "nb_floral_unit_zero_rows": sum(value == 0 for value in finite),
        "nb_floral_unit_positive_rows": sum(value > 0 for value in finite),
        "missing_floral_unit_rows_with_nonblank_plant_best_id": sum(bool(clean(row.get("Plant_Best_ID"))) for row in missing_rows),
        "missing_floral_unit_rows_with_nonblank_name_floral_unit": sum(bool(clean(row.get("Name_Floral_unit"))) for row in missing_rows),
        "missing_floral_unit_rows_context_count": len({
            (clean(row.get("Site")), clean(row.get("Period")))
            for row in missing_rows if clean(row.get("Site")) and clean(row.get("Period"))
        }),
        "missing_floral_unit_rows_preview": missing_preview,
        "site_periods_with_quadrat_ids": len(quadrat_by_context),
        "quadrat_count_per_site_period_distinct": sorted({len(values) for values in quadrat_by_context.values()}),
        "binary_opportunity_boundary": (
            "This audit does not yet decide whether a floral row with missing Nb_Floral_unit is an active plant. "
            "That rule is frozen only after the missing-row structure is inspected."
        ),
    }


def main() -> None:
    payloads = fetch_required_bytes()
    interaction_headers, interaction_rows = rows_from_sheet(payloads[INTERACTION_NAME], "Insects-Plants")
    _sampling_interaction_headers, sampling_interaction_rows = rows_from_sheet(payloads[SAMPLING_NAME], "Insects_Plants")
    floral_headers, floral_rows = rows_from_sheet(payloads[SAMPLING_NAME], "Floral_abundance")

    interaction_context = context_keys(interaction_rows)
    sampling_interaction_context = context_keys(sampling_interaction_rows)
    floral_context = context_keys(floral_rows)
    interaction_period_month = period_month_map(interaction_rows)
    floral_period_month = period_month_map(floral_rows)
    exact_period_month_mapping = interaction_period_month == floral_period_month and all(len(v) == 1 for v in interaction_period_month.values())

    output = {
        "schema_version": "1.1",
        "analysis": "martinique_2025_reconstruction_structure_audit",
        "target_metrics_calculated": False,
        "network_matrices_built": False,
        "interaction_events_aggregated": False,
        "v9_predictive_fit_calculated": False,
        "source_hashes": {name: source_gate.sha256(payload) for name, payload in payloads.items()},
        "interaction": {
            "headers": interaction_headers,
            "row_count": len(interaction_rows),
            "context_structure": interaction_context,
            "period_to_months": interaction_period_month,
            "site_coverage_by_month": site_month_coverage(interaction_rows),
            "plant_identity": identity_structure(interaction_rows, "Plant_Best_ID", "Plant_genus", "Plant_species", "Plant_family"),
            "insect_identity": identity_structure(interaction_rows, "Insect_Best_ID", "Insect_genus", "Insect_species", "Insect_family"),
            "joint_identity_structure": joint_interaction_identity_structure(interaction_rows),
            "timing_field_structure": timing_field_structure(interaction_rows),
        },
        "sampling_workbook_interaction_sheet": {
            "row_count": len(sampling_interaction_rows),
            "same_row_count_as_former_names_workbook": len(sampling_interaction_rows) == len(interaction_rows),
            "context_structure": sampling_interaction_context,
        },
        "floral_abundance": {
            "headers": floral_headers,
            "row_count": len(floral_rows),
            "context_structure": floral_context,
            "period_to_months": floral_period_month,
            "site_coverage_by_month": site_month_coverage(floral_rows),
            "plant_identity": identity_structure(floral_rows, "Plant_Best_ID", "Plant_genus", "Plant_species", "Plant_family"),
            "floral_measure_structure": floral_structure(floral_rows),
        },
        "cross_source_structure": {
            "site_sets_match": set(interaction_context["sites"]) == set(floral_context["sites"]),
            "month_sets_match": set(interaction_context["months"]) == set(floral_context["months"]),
            "period_sets_match": set(interaction_context["periods"]) == set(floral_context["periods"]),
            "period_is_one_to_one_with_month_in_both_sources": exact_period_month_mapping,
            "interaction_site_period_keys_equal_floral": {
                tuple(row) for row in interaction_context["site_period_keys"]
            } == {tuple(row) for row in floral_context["site_period_keys"]},
            "complete_10x12_site_period_grid_in_both_sources": (
                interaction_context["site_count"] == 10 and interaction_context["period_count"] == 12
                and interaction_context["site_period_count"] == 120
                and floral_context["site_period_count"] == 120
            ),
        },
        "claim_boundary": (
            "Source-structure audit only. Context identity, joint blank patterns, event timing fields, and missing floral-unit records are inspected before selecting network units or calculating any ecological network target."
        ),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(output, indent=2, ensure_ascii=False) + "\n")
    print(json.dumps({
        "interaction_rows": len(interaction_rows),
        "floral_rows": len(floral_rows),
        "interaction_context": interaction_context,
        "floral_context": floral_context,
        "period_month_mapping": interaction_period_month,
        "joint_identity_structure": output["interaction"]["joint_identity_structure"],
        "timing_field_structure": output["interaction"]["timing_field_structure"],
        "floral_measure_structure": output["floral_abundance"]["floral_measure_structure"],
        "cross_source_structure": output["cross_source_structure"],
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
