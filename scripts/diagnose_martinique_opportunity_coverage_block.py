from __future__ import annotations

import io
import json
import hashlib
import urllib.request
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "data/results/martinique_opportunity_coverage_diagnosis.json"
FORMER_URL = "https://search-data.ubfc.fr/dl_data.php?file=597"
SAMPLING_URL = "https://search-data.ubfc.fr/dl_data.php?file=601"
FORMER_SHA = "9a001287bf64d51cbdefee1164579398e0cf5053efbfc04d1f8bcf9338626753"
SAMPLING_SHA = "e3f82dc81749d7c759dbb62fc2e40ceeff9382758a3114c63c57553d15c2327d"
USER_AGENT = "izu-core-martinique-opportunity-diagnosis/1.0"
MISSING = {"", "na", "n/a", "nan", "none", "null"}


def fetch(url: str, expected_sha: str) -> bytes:
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=120) as response:
        payload = response.read()
    actual = hashlib.sha256(payload).hexdigest()
    if actual != expected_sha:
        raise RuntimeError(f"source checksum drift: {actual}")
    return payload


def rows(payload: bytes, sheet_name: str) -> list[dict[str, object]]:
    book = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    sheet = book[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(v).strip() if v is not None else "" for v in next(iterator)]
    result = []
    for raw in iterator:
        if not any(v not in (None, "") for v in raw):
            continue
        result.append({h: v for h, v in zip(headers, raw) if h})
    book.close()
    return result


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def identity(value: object) -> str:
    text = clean(value)
    return "" if text.casefold() in MISSING else text


def context(row: dict[str, object]) -> tuple[str, str]:
    return clean(row.get("Site")), clean(row.get("Period"))


def structural_key(row: dict[str, object]) -> tuple[str, ...]:
    return tuple(clean(row.get(field)) for field in (
        "Period", "Date", "Site", "Transect", "H_start", "H_end", "Num_sp"
    ))


def main() -> None:
    former_payload = fetch(FORMER_URL, FORMER_SHA)
    sampling_payload = fetch(SAMPLING_URL, SAMPLING_SHA)
    former = rows(former_payload, "Insects-Plants")
    current = rows(sampling_payload, "Insects_Plants")
    floral = rows(sampling_payload, "Floral_abundance")
    if len(former) != 4553 or len(current) != 4553 or len(floral) != 8546:
        raise RuntimeError("Martinique source row count drift")

    floral_by_context: dict[tuple[str, str], set[str]] = defaultdict(set)
    floral_by_site: dict[str, set[str]] = defaultdict(set)
    floral_by_period: dict[str, set[str]] = defaultdict(set)
    floral_global: set[str] = set()
    quadrats_by_context: dict[tuple[str, str], set[tuple[str, str]]] = defaultdict(set)
    for row in floral:
        plant = identity(row.get("Plant_Best_ID"))
        ctx = context(row)
        if plant:
            floral_by_context[ctx].add(plant)
            floral_by_site[ctx[0]].add(plant)
            floral_by_period[ctx[1]].add(plant)
            floral_global.add(plant)
        quadrat = identity(row.get("Quadrat"))
        transect = identity(row.get("Transect"))
        if ctx[0] and ctx[1] and quadrat:
            quadrats_by_context[ctx].add((transect, quadrat))

    former_by_key: dict[tuple[str, ...], list[dict[str, object]]] = defaultdict(list)
    current_by_key: dict[tuple[str, ...], list[dict[str, object]]] = defaultdict(list)
    for row in former:
        former_by_key[structural_key(row)].append(row)
    for row in current:
        current_by_key[structural_key(row)].append(row)

    all_keys = set(former_by_key) | set(current_by_key)
    row_alignment = {
        "former_structural_key_count": len(former_by_key),
        "current_structural_key_count": len(current_by_key),
        "structural_key_sets_equal": set(former_by_key) == set(current_by_key),
        "keys_with_count_mismatch": sum(len(former_by_key[key]) != len(current_by_key[key]) for key in all_keys),
    }

    former_interaction_plants: dict[tuple[str, str], set[str]] = defaultdict(set)
    current_interaction_plants: dict[tuple[str, str], set[str]] = defaultdict(set)
    for row in former:
        plant = identity(row.get("Plant_Best_ID"))
        insect = identity(row.get("Insect_Best_ID"))
        if plant and insect:
            former_interaction_plants[context(row)].add(plant)
    for row in current:
        plant = identity(row.get("Plant_Best_ID"))
        insect = identity(row.get("Insect_Best_ID"))
        if plant and insect:
            current_interaction_plants[context(row)].add(plant)

    mismatch_rows = []
    category_counts = Counter()
    former_missing_global = set()
    current_missing_global = set()
    contexts = sorted(set(former_interaction_plants) | set(floral_by_context))
    for ctx in contexts:
        former_missing = sorted(former_interaction_plants[ctx] - floral_by_context[ctx])
        current_missing = sorted(current_interaction_plants[ctx] - floral_by_context[ctx])
        former_only_mismatch = set(former_missing)
        current_only_mismatch = set(current_missing)
        for plant in former_missing:
            if plant not in floral_global:
                category = "former_name_absent_from_floral_globally"
                former_missing_global.add(plant)
            elif plant in floral_by_site[ctx[0]]:
                category = "present_in_floral_same_site_other_period"
            elif plant in floral_by_period[ctx[1]]:
                category = "present_in_floral_same_period_other_site"
            else:
                category = "present_in_floral_other_site_and_period"
            category_counts[category] += 1
        for plant in current_missing:
            if plant not in floral_global:
                current_missing_global.add(plant)
        mismatch_rows.append({
            "site": ctx[0],
            "period": ctx[1],
            "former_interaction_plant_count": len(former_interaction_plants[ctx]),
            "current_interaction_plant_count": len(current_interaction_plants[ctx]),
            "floral_plant_count": len(floral_by_context[ctx]),
            "former_missing_count": len(former_missing),
            "current_missing_count": len(current_missing),
            "former_missing": former_missing,
            "current_missing": current_missing,
            "former_only_vs_current_mismatch": sorted(former_only_mismatch - current_only_mismatch),
            "current_only_vs_former_mismatch": sorted(current_only_mismatch - former_only_mismatch),
            "quadrat_count": len(quadrats_by_context[ctx]),
        })

    former_total = sum(row["former_missing_count"] for row in mismatch_rows)
    current_total = sum(row["current_missing_count"] for row in mismatch_rows)
    former_contexts = sum(row["former_missing_count"] > 0 for row in mismatch_rows)
    current_contexts = sum(row["current_missing_count"] > 0 for row in mismatch_rows)

    output = {
        "schema_version": "1.0",
        "analysis": "martinique_opportunity_coverage_post_block_diagnosis",
        "preserved_block_decision": "blocked_martinique_independent_plant_opportunity_does_not_cover_observed_interaction_plants",
        "v9_parameters_reestimated": False,
        "primary_validation_relabelled": False,
        "target_metrics_calculated": False,
        "network_outcomes_inspected": False,
        "source_hashes": {
            "former_interaction": FORMER_SHA,
            "sampling_workbook": SAMPLING_SHA,
        },
        "row_alignment": row_alignment,
        "coverage_summary": {
            "context_count": len(contexts),
            "former_name_mismatch_context_count": former_contexts,
            "former_name_mismatch_endpoint_count": former_total,
            "current_name_mismatch_context_count": current_contexts,
            "current_name_mismatch_endpoint_count": current_total,
            "endpoint_mismatch_reduction_using_current_interaction_names": former_total - current_total,
            "context_mismatch_reduction_using_current_interaction_names": former_contexts - current_contexts,
            "former_name_mismatch_categories": dict(category_counts),
            "former_interaction_plants_absent_from_floral_global_count": len(former_missing_global),
            "current_interaction_plants_absent_from_floral_global_count": len(current_missing_global),
            "former_interaction_plants_absent_from_floral_global": sorted(former_missing_global),
            "current_interaction_plants_absent_from_floral_global": sorted(current_missing_global),
            "quadrat_count_per_context_distinct": sorted({len(value) for value in quadrats_by_context.values()}),
        },
        "contexts": mismatch_rows,
        "interpretation_boundary": (
            "Post-block source/measurement diagnosis only. Comparing the deposited former-name interaction workbook with the current-name interaction sheet may identify taxonomy-version contribution, while same-site/other-period and same-period/other-site floral presence diagnoses sampling coverage. None of these comparisons rescues PR #213 or permits Martinique to become a clean v9 confirmation."
        ),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(output, indent=2, ensure_ascii=False) + "\n")
    print(json.dumps({
        "row_alignment": row_alignment,
        "coverage_summary": output["coverage_summary"],
    }, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
