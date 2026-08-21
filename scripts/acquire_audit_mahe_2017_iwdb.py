from __future__ import annotations

import hashlib
import io
import json
import re
import urllib.error
import urllib.parse
import urllib.request
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DESIGN = ROOT / "data/design/abm_v9_mahe_iwdb_source_gate_v1.json"
OUT = ROOT / "data/results/mahe_2017_iwdb_source_audit.json"
RAW_DIR = ROOT / "data/external/mahe_2017_iwdb"
USER_AGENT = "izu-core-mahe-source-audit/1.0"


class LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.links: list[tuple[str, str]] = []
        self._href: str | None = None
        self._text: list[str] = []

    def handle_starttag(self, tag: str, attrs) -> None:
        if tag.lower() != "a":
            return
        values = dict(attrs)
        self._href = values.get("href")
        self._text = []

    def handle_data(self, data: str) -> None:
        if self._href is not None:
            self._text.append(data)

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() == "a" and self._href is not None:
            self.links.append((self._href, " ".join(" ".join(self._text).split())))
            self._href = None
            self._text = []


def fetch_bytes(url: str) -> tuple[int | None, bytes | None, str | None, str | None]:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=90) as response:
            return int(response.status), response.read(), None, response.geturl()
    except urllib.error.HTTPError as exc:
        return int(exc.code), None, str(exc), None
    except Exception as exc:
        return None, None, f"{type(exc).__name__}: {exc}", None


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def candidate_header(preview: list[list[object]]) -> tuple[int | None, list[str]]:
    best_row = None
    best_headers: list[str] = []
    best_score = -1
    tokens = ("treatment", "site", "month", "network", "plant", "floral abundance")
    for index, row in enumerate(preview, start=1):
        headers = [str(value).strip() if value is not None else "" for value in row]
        normalized = [normalize(value) for value in headers]
        score = sum(any(token in value for value in normalized) for token in tokens)
        if score > best_score:
            best_score = score
            best_row = index
            best_headers = headers
    if best_score < 3:
        return None, []
    return best_row, best_headers


def metadata_column_indices(headers: list[str]) -> dict[str, int | None]:
    normalized = [normalize(value) for value in headers]
    rules = {
        "treatment": ("treatment",),
        "site": ("site",),
        "month": ("month",),
        "network_id": ("network id", "network code", "network"),
        "plant_id": ("plant species id", "plant species", "plant id"),
        "floral_abundance": ("floral abundance", "flower abundance"),
    }
    result: dict[str, int | None] = {}
    for role, tokens in rules.items():
        matches = [
            index for index, value in enumerate(normalized)
            if any(token == value or token in value for token in tokens)
        ]
        result[role] = matches[0] if matches else None
    return result


def summarize_metadata_rows(rows: list[list[object]], header_row: int, headers: list[str]) -> dict:
    indices = metadata_column_indices(headers)
    values: dict[str, set[str]] = {key: set() for key in indices}
    nonblank_data_rows = 0
    for row in rows[header_row:]:
        if not any(value not in (None, "") for value in row):
            continue
        nonblank_data_rows += 1
        for role, index in indices.items():
            if index is None or index >= len(row):
                continue
            value = row[index]
            text = " ".join(str(value).split()) if value is not None else ""
            if text:
                values[role].add(text)
    metadata_indices = {index for index in indices.values() if index is not None}
    interaction_column_count = sum(
        bool(str(header).strip()) and index not in metadata_indices
        for index, header in enumerate(headers)
    )
    return {
        "metadata_column_indices_zero_based": indices,
        "nonblank_data_row_count": nonblank_data_rows,
        "unique_counts": {role: len(items) for role, items in values.items()},
        "unique_network_ids": sorted(values["network_id"]),
        "sites": sorted(values["site"]),
        "months": sorted(values["month"]),
        "treatments": sorted(values["treatment"]),
        "interaction_column_count_excluding_recognized_metadata": interaction_column_count,
    }


def inventory_xlsx(payload: bytes) -> dict:
    book = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    sheets = []
    for sheet in book.worksheets:
        rows = [list(row) for row in sheet.iter_rows(values_only=True)]
        preview = rows[:15]
        header_row, headers = candidate_header(preview)
        entry = {
            "name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "candidate_header_row": header_row,
            "candidate_headers": headers,
        }
        if header_row is not None:
            entry["metadata_structure"] = summarize_metadata_rows(rows, header_row, headers)
        sheets.append(entry)
    book.close()
    return {"format": "xlsx", "sheet_count": len(sheets), "sheets": sheets}


def inventory_xls(payload: bytes) -> dict:
    try:
        import xlrd  # type: ignore
    except ImportError as exc:
        raise RuntimeError("xlrd is required for .xls source audit") from exc
    book = xlrd.open_workbook(file_contents=payload)
    sheets = []
    for sheet in book.sheets():
        rows = [sheet.row_values(index) for index in range(sheet.nrows)]
        preview = rows[:15]
        header_row, headers = candidate_header(preview)
        entry = {
            "name": sheet.name,
            "max_row": sheet.nrows,
            "max_column": sheet.ncols,
            "candidate_header_row": header_row,
            "candidate_headers": headers,
        }
        if header_row is not None:
            entry["metadata_structure"] = summarize_metadata_rows(rows, header_row, headers)
        sheets.append(entry)
    return {"format": "xls", "sheet_count": len(sheets), "sheets": sheets}


def workbook_has_64_network_matrix(inventory: dict) -> bool:
    for sheet in inventory.get("sheets", []):
        structure = sheet.get("metadata_structure") or {}
        counts = structure.get("unique_counts") or {}
        indices = structure.get("metadata_column_indices_zero_based") or {}
        required = ("site", "month", "network_id", "plant_id", "floral_abundance")
        if (
            counts.get("network_id") == 64
            and all(indices.get(role) is not None for role in required)
            and structure.get("interaction_column_count_excluding_recognized_metadata", 0) > 0
        ):
            return True
    return False


def representation_labels(file_records: list[dict]) -> dict[str, bool]:
    texts = []
    for row in file_records:
        texts.append(str(row.get("name", "")).lower())
        inventory = row.get("workbook_inventory") or {}
        texts.extend(str(sheet.get("name", "")).lower() for sheet in inventory.get("sheets", []))
    joined = " ".join(texts)
    return {
        "no_visits_visible": "no.visits" in joined or "no visits" in joined or "visits" in joined,
        "visitfreq_visible": "visitfreq" in joined or "visit freq" in joined or "visitation frequency" in joined,
    }


def write(payload: dict) -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
    print(json.dumps(payload, indent=2, ensure_ascii=False))


def main() -> None:
    design = json.loads(DESIGN.read_text())
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    page_attempts = []
    page_url = None
    page_payload = None
    resolved_page_url = None
    for url in design["candidate_system"]["database_entry_paths"]:
        status, payload, error, resolved = fetch_bytes(url)
        page_attempts.append({"url": url, "http_status": status, "error": error, "resolved_url": resolved})
        if status == 200 and payload:
            page_url = url
            page_payload = payload
            resolved_page_url = resolved or url
            break

    if page_payload is None or resolved_page_url is None:
        write({
            "schema_version": "1.0",
            "analysis": "mahe_2017_iwdb_source_audit",
            "status": "blocked_mahe_iwdb_entry_not_recovered",
            "page_attempts": page_attempts,
            "source_admission_succeeds": False,
            "target_metrics_calculated": False,
        })
        return

    page_sha = hashlib.sha256(page_payload).hexdigest()
    page_text = page_payload.decode("utf-8", errors="replace")
    parser = LinkParser()
    parser.feed(page_text)
    candidate_links = []
    seen = set()
    for href, anchor in parser.links:
        absolute = urllib.parse.urljoin(resolved_page_url, href)
        path = urllib.parse.urlparse(absolute).path.lower()
        if path.endswith((".xlsx", ".xls")) and absolute not in seen:
            seen.add(absolute)
            candidate_links.append({"href": href, "anchor": anchor, "url": absolute})

    file_records = []
    blocked_files = []
    for link in candidate_links:
        status, payload, error, resolved = fetch_bytes(link["url"])
        name = Path(urllib.parse.urlparse(resolved or link["url"]).path).name or Path(urllib.parse.urlparse(link["url"]).path).name
        record = {
            **link,
            "name": urllib.parse.unquote(name),
            "http_status": status,
            "error": error,
            "resolved_url": resolved,
        }
        if status != 200 or payload is None:
            blocked_files.append(record["name"] or link["url"])
            file_records.append(record)
            continue
        record.update({
            "bytes": len(payload),
            "md5": hashlib.md5(payload).hexdigest(),
            "sha256": hashlib.sha256(payload).hexdigest(),
        })
        safe_name = re.sub(r"[^A-Za-z0-9._-]+", "_", record["name"] or f"iwdb_{len(file_records)}.xlsx")
        (RAW_DIR / safe_name).write_bytes(payload)
        suffix = Path(record["name"]).suffix.lower()
        try:
            if suffix == ".xlsx":
                record["workbook_inventory"] = inventory_xlsx(payload)
            elif suffix == ".xls":
                record["workbook_inventory"] = inventory_xls(payload)
            else:
                record["inventory_error"] = "unrecognized Excel suffix"
        except Exception as exc:
            record["inventory_error"] = f"{type(exc).__name__}: {exc}"
        file_records.append(record)

    source_bytes_ok = bool(candidate_links) and not blocked_files and all(row.get("bytes") for row in file_records)
    matrix_64_visible = any(
        workbook_has_64_network_matrix(row.get("workbook_inventory") or {})
        for row in file_records
    )
    representations = representation_labels(file_records)
    admission = source_bytes_ok and matrix_64_visible

    status_name = (
        "source_admitted_mahe_iwdb_64_raw_monthly_matrices"
        if admission
        else "blocked_mahe_iwdb_raw_matrix_bytes_or_64_network_schema_incomplete"
    )
    write({
        "schema_version": "1.0",
        "analysis": "mahe_2017_iwdb_source_audit",
        "status": status_name,
        "database_entry_requested": page_url,
        "database_entry_resolved": resolved_page_url,
        "page_attempts": page_attempts,
        "database_entry_bytes": len(page_payload),
        "database_entry_sha256": page_sha,
        "excel_link_count": len(candidate_links),
        "excel_links": candidate_links,
        "files": file_records,
        "blocked_files": blocked_files,
        "source_bytes_ok": source_bytes_ok,
        "raw_64_network_matrix_structure_visible": matrix_64_visible,
        "representation_labels": representations,
        "source_admission_succeeds": admission,
        "target_metrics_calculated": False,
        "prior_block_boundary": design["prior_block_boundary"],
        "observation_boundary": design["source_only_gate"]["observation_boundary"],
        "claim_boundary": (
            "Source transport/schema audit only. The database entry and its directly linked raw Excel matrices are inspected for provenance and repeated-network structure. "
            "No pair-support, diversity, overlap, local-range, or ABM v9 predictive statistic is calculated."
        ),
    })


if __name__ == "__main__":
    main()
