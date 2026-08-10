#!/usr/bin/env python3
"""Analyse the source-native 2026 Ogasawara interaction workbook."""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Mapping, Sequence

from channel_id.ogasawara_context_network import (
    analyze_anijima_anole_contrast,
    context_metadata,
    context_season_metrics,
    validate_rows,
)


SOURCE_COLUMNS = {
    "Island": "island",
    "Invasional Context": "context",
    "Season": "season",
    "Forest_Status": "forest_status",
    "Anole": "anole",
    "Plant_sp": "plant",
    "Poll_sp": "pollinator",
    "N.Int": "interaction_count",
}


def _write_csv(path: Path, rows: Sequence[Mapping[str, object]]) -> None:
    columns: list[str] = []
    for row in rows:
        for key in row:
            if key not in columns:
                columns.append(key)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    key: (
                        json.dumps(
                            value,
                            ensure_ascii=False,
                            separators=(",", ":"),
                        )
                        if isinstance(value, (list, tuple, dict))
                        else value
                    )
                    for key, value in row.items()
                }
            )


def locate_workbook(
    source_dir: Path, inventory: Mapping[str, object]
) -> tuple[Path, str]:
    candidates: list[tuple[Path, str]] = []
    for record in inventory.get("files", []):
        if not isinstance(record, dict) or record.get("status") != "downloaded":
            continue
        local_name = str(record.get("local_name") or record.get("key") or "")
        if local_name.lower().endswith((".xlsx", ".xlsm")):
            candidates.append(
                (
                    source_dir / "files" / local_name,
                    str(record.get("sha256") or ""),
                )
            )
    candidates = [
        (path, digest) for path, digest in candidates if path.exists()
    ]
    if len(candidates) != 1:
        raise ValueError(
            "expected exactly one downloaded Ogasawara workbook, "
            f"found {len(candidates)}"
        )
    return candidates[0]


def read_source_rows(workbook_path: Path) -> tuple[dict[str, object], ...]:
    import openpyxl

    workbook = openpyxl.load_workbook(
        workbook_path, read_only=True, data_only=True
    )
    if "Data interactions" not in workbook.sheetnames:
        raise ValueError(
            "source workbook lacks the expected 'Data interactions' sheet"
        )
    worksheet = workbook["Data interactions"]
    iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value or "").strip() for value in next(iterator)]
    missing = set(SOURCE_COLUMNS) - set(headers)
    if missing:
        raise ValueError(
            "source workbook missing columns: " + ", ".join(sorted(missing))
        )
    index = {header: headers.index(header) for header in SOURCE_COLUMNS}
    rows: list[dict[str, object]] = []
    for source_row in iterator:
        if not any(value not in (None, "") for value in source_row):
            continue
        rows.append(
            {
                target: (
                    source_row[index[source]]
                    if index[source] < len(source_row)
                    else None
                )
                for source, target in SOURCE_COLUMNS.items()
            }
        )
    workbook.close()
    return tuple(rows)


def build_report(analysis: Mapping[str, object]) -> str:
    contrast = analysis["anijima_anole_contrast"]
    effects = {
        row["effect_id"]: row
        for row in contrast["effect_level_uncertainty"]["effects"]
    }
    return "\n".join(
        [
            "# Ogasawara source-native context-network analysis",
            "",
            "## Source design",
            "",
            f"- legitimate-interaction rows: **{analysis['n_source_rows']}**",
            f"- source-defined invasion contexts: **{analysis['n_contexts']}**",
            (
                "- context × season networks: "
                f"**{analysis['n_context_season_networks']}**"
            ),
            "",
            "## Within-Anijima natural-forest contrast",
            "",
            (
                "- shared plant × season cells: "
                f"**{contrast['n_plant_season_contrasts']}**"
            ),
            (
                "- unique shared plant effect units: "
                f"**{contrast['n_unique_shared_plants']}**"
            ),
            "",
            "| effect | median | plant-bootstrap 95% interval |",
            "|---|---:|---:|",
            (
                "| legitimate-interaction LRR, presence/absence | "
                f"{effects['ogasawara_anijima_visitation_lrr']['estimate']:.3f} | "
                f"[{effects['ogasawara_anijima_visitation_lrr']['uncertainty_value'][0]:.3f}, "
                f"{effects['ogasawara_anijima_visitation_lrr']['uncertainty_value'][1]:.3f}] |"
            ),
            (
                "| pollinator-richness LRR, presence/absence | "
                f"{effects['ogasawara_anijima_pollinator_richness_lrr']['estimate']:.3f} | "
                f"[{effects['ogasawara_anijima_pollinator_richness_lrr']['uncertainty_value'][0]:.3f}, "
                f"{effects['ogasawara_anijima_pollinator_richness_lrr']['uncertainty_value'][1]:.3f}] |"
            ),
            (
                "| pollinator-assemblage turnover | "
                f"{effects['ogasawara_anijima_partner_turnover']['estimate']:.3f} | "
                f"[{effects['ogasawara_anijima_partner_turnover']['uncertainty_value'][0]:.3f}, "
                f"{effects['ogasawara_anijima_partner_turnover']['uncertainty_value'][1]:.3f}] |"
            ),
            "",
            "## Claim boundary",
            "",
            str(analysis["claim_boundary"]),
            "",
        ]
    )


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source-dir",
        type=Path,
        default=Path("artifacts/ogasawara_zenodo"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("artifacts/ogasawara_context_analysis"),
    )
    args = parser.parse_args()

    inventory_path = args.source_dir / "source_inventory.json"
    inventory = json.loads(inventory_path.read_text(encoding="utf-8"))
    if inventory.get("status") != "acquired":
        raise ValueError("Ogasawara source inventory is not in acquired state")
    workbook_path, source_sha256 = locate_workbook(args.source_dir, inventory)
    rows = validate_rows(read_source_rows(workbook_path))
    metrics = context_season_metrics(rows)
    anijima = analyze_anijima_anole_contrast(
        rows, source_sha256=source_sha256
    )

    analysis = {
        "schema_version": "1.0",
        "status": "source_resolved_context_network_analysis",
        "source_id": inventory.get("source_id"),
        "dataset_doi": inventory.get("dataset_doi"),
        "license": inventory.get("license"),
        "source_workbook": workbook_path.name,
        "source_sha256": source_sha256,
        "n_source_rows": len(rows),
        "n_zero_marker_rows": sum(
            float(row["interaction_count"]) == 0 for row in rows
        ),
        "n_contexts": len(context_metadata(rows)),
        "n_context_season_networks": len(metrics),
        "context_metadata": context_metadata(rows),
        "context_season_metrics": metrics,
        "anijima_anole_contrast": anijima,
        "methods": {
            "interaction_semantics": (
                "Source N.Int legitimate-contact counts; No_pollinator rows with "
                "zero count are retained as source markers but excluded from "
                "weighted partner vectors."
            ),
            "context_networks": (
                "Each source-defined invasion context and season is summarized "
                "separately before any contrast."
            ),
            "matched_contrast": (
                "Within Anijima natural forest, plants are matched between "
                "source-defined anole-presence and anole-absence contexts within "
                "season. Repeated seasons are collapsed to a within-plant median "
                "before exact plant bootstrap."
            ),
        },
        "claim_boundary": (
            "The analysis quantifies legitimate-contact network structure and a "
            "spatially structured Anijima anole-context contrast. Contexts are not "
            "randomized, source rows lack site identifiers, and plant bootstrap "
            "intervals do not create independent invasion contexts or archipelagos. "
            "Legitimate interaction counts are not pollen deposition, pollinator "
            "effectiveness, reproductive success, effective dependency, or "
            "historical mainland-distance effects."
        ),
    }

    args.output_dir.mkdir(parents=True, exist_ok=True)
    (args.output_dir / "analysis.json").write_text(
        json.dumps(analysis, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    effect_document = anijima["effect_level_uncertainty"]
    (args.output_dir / "effect_rows.json").write_text(
        json.dumps(effect_document, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    _write_csv(
        args.output_dir / "context_season_metrics.csv", list(metrics)
    )
    _write_csv(
        args.output_dir / "anijima_plant_season_contrasts.csv",
        anijima["plant_season_contrasts"],
    )
    _write_csv(
        args.output_dir / "anijima_plant_effect_units.csv",
        anijima["plant_level_effect_units"],
    )
    (args.output_dir / "report.md").write_text(
        build_report(analysis), encoding="utf-8"
    )
    print(f"source rows: {analysis['n_source_rows']}")
    print(f"contexts: {analysis['n_contexts']}")
    print(f"matched plants: {anijima['n_unique_shared_plants']}")
    print(args.output_dir / "analysis.json")


if __name__ == "__main__":
    main()
