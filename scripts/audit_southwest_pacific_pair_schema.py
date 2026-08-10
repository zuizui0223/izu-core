#!/usr/bin/env python3
"""Audit source-native mainland–island floral comparison tables."""
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Iterable, Sequence


ALIASES = {
    "comparison_id": ("comparison_id", "comparison", "pair_id", "pair", "colonisation", "colonization", "event_id"),
    "archipelago": ("archipelago", "island_group", "island_system"),
    "island": ("island", "island_name"),
    "island_taxon": ("island_taxon", "island_species", "insular_taxon", "insular_species"),
    "mainland_taxon": ("mainland_taxon", "mainland_species", "continental_taxon", "source_taxon", "continental_species"),
    "trait": ("trait", "floral_trait", "flower_trait", "measurement"),
    "unit": ("unit", "units"),
    "island_value": ("island_value", "insular_value", "island_flower", "insular_flower"),
    "mainland_value": ("mainland_value", "continental_value", "source_value", "mainland_flower", "continental_flower"),
    "group": ("realm", "origin_group", "island_mainland", "population_type", "geographic_group"),
    "value": ("value", "trait_value", "measurement_value", "mean"),
    "pollination_mode": ("pollination_mode", "pollination_system", "pollination", "animal_pollinated", "wind_pollinated"),
    "breeding_system": ("breeding_system", "mating_system", "self_compatibility", "reproductive_system"),
    "island_type": ("island_type", "geological_origin", "geological_type", "oceanic_continental"),
    "growth_form": ("growth_form", "life_form", "habit"),
    "sample_size": ("sample_size", "n_island", "n_mainland", "number_of_individuals"),
    "uncertainty": ("sd", "se", "standard_deviation", "standard_error", "confidence_interval"),
}


def norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().casefold()).strip("_")


def roles(headers: Sequence[object]) -> dict[str, list[str]]:
    normalized = [(norm(header), str(header or "")) for header in headers if str(header or "").strip()]
    output = {}
    for role, aliases in ALIASES.items():
        output[role] = [
            original
            for key, original in normalized
            if any(key == alias or alias in key for alias in aliases)
        ]
    return output


def first_block(rows: Iterable[Sequence[object]]) -> tuple[list[object], list[list[object]]]:
    iterator = iter(rows)
    headers = []
    for row in iterator:
        values = list(row)
        if any(str(value or "").strip() for value in values):
            headers = values
            break
    sample = []
    for row in iterator:
        values = list(row)
        if any(str(value or "").strip() for value in values):
            sample.append(values)
        if len(sample) >= 20:
            break
    return headers, sample


def classify(
    path: Path,
    sheet: str | None,
    headers: list[object],
    sample: list[list[object]],
    n_rows: int,
    n_columns: int,
    delimiter: str | None,
) -> dict[str, object]:
    resolved = roles(headers)
    matched = {role for role, values in resolved.items() if values}
    paired_wide = (
        {"island_value", "mainland_value"}.issubset(matched)
        and bool(matched & {"comparison_id", "island_taxon", "mainland_taxon"})
    )
    paired_long = (
        {"group", "value"}.issubset(matched)
        and bool(matched & {"comparison_id", "island_taxon", "mainland_taxon"})
    )
    moderator_table = bool(matched & {"pollination_mode", "breeding_system", "island_type", "growth_form"})
    quantitative_ready = (paired_wide or paired_long) and bool(matched & {"trait", "unit"})
    return {
        "file": str(path),
        "sheet": sheet,
        "n_rows_reported": n_rows,
        "n_columns_reported": n_columns,
        "delimiter": delimiter,
        "headers": [str(value or "") for value in headers],
        "role_matches": resolved,
        "paired_wide_candidate": paired_wide,
        "paired_long_candidate": paired_long,
        "moderator_table_candidate": moderator_table,
        "quantitative_pair_candidate": quantitative_ready,
        "contains_sample_size_field": "sample_size" in matched,
        "contains_uncertainty_field": "uncertainty" in matched,
        "sample_preview": sample[:5],
    }


def read_delimited(path: Path) -> list[dict[str, object]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    delimiter = "\t" if path.suffix.casefold() == ".tsv" else ","
    try:
        delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",\t;").delimiter
    except csv.Error:
        pass
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    headers, sample = first_block(rows)
    return [classify(path, None, headers, sample, len(rows), max((len(row) for row in rows), default=0), delimiter)]


def read_xlsx(path: Path) -> list[dict[str, object]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    output = []
    for worksheet in workbook.worksheets:
        headers, sample = first_block(worksheet.iter_rows(values_only=True))
        output.append(classify(path, worksheet.title, headers, sample, worksheet.max_row, worksheet.max_column, None))
    workbook.close()
    return output


def read_docx(path: Path) -> list[dict[str, object]]:
    from docx import Document

    document = Document(path)
    output = []
    for index, table in enumerate(document.tables, start=1):
        rows = [[cell.text for cell in row.cells] for row in table.rows]
        headers, sample = first_block(rows)
        output.append(
            classify(
                path,
                f"docx_table_{index}",
                headers,
                sample,
                len(rows),
                max((len(row) for row in rows), default=0),
                None,
            )
        )
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", type=Path, default=Path("artifacts/southwest_pacific_aob"))
    parser.add_argument("--output", type=Path, default=Path("artifacts/southwest_pacific_aob/schema_audit.json"))
    args = parser.parse_args()

    tables = []
    errors = []
    for path in sorted(args.input_dir.rglob("*")):
        if not path.is_file() or path.name in {"source_inventory.json", "schema_audit.json"}:
            continue
        try:
            suffix = path.suffix.casefold()
            if suffix in {".csv", ".tsv", ".txt"}:
                tables.extend(read_delimited(path))
            elif suffix in {".xlsx", ".xlsm"}:
                tables.extend(read_xlsx(path))
            elif suffix == ".docx":
                tables.extend(read_docx(path))
        except Exception as error:
            errors.append({"file": str(path), "error": repr(error)})

    summary = {
        "status": "tabular_schema_audited" if tables else "no_parseable_tabular_supplement",
        "n_tabular_units": len(tables),
        "n_paired_wide_candidates": sum(bool(row["paired_wide_candidate"]) for row in tables),
        "n_paired_long_candidates": sum(bool(row["paired_long_candidate"]) for row in tables),
        "n_quantitative_pair_candidates": sum(bool(row["quantitative_pair_candidate"]) for row in tables),
        "n_moderator_table_candidates": sum(bool(row["moderator_table_candidate"]) for row in tables),
        "tables": tables,
        "errors": errors,
        "next_gate": (
            "Verify the source-defined comparison identifier, island/mainland orientation, trait units and shared sampling hierarchy. Only then compute within-comparison response sizes."
        ),
        "claim_boundary": (
            "Header matching does not establish 129 independent biological replicates, pollinator dependency or causal island evolution. Pair-level morphology may be analysed only after comparison identity, orientation and trait definition are source-resolved."
        ),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(summary, indent=2, ensure_ascii=False, default=str) + "\n", encoding="utf-8")
    print(f"tabular units: {summary['n_tabular_units']}")
    print(f"quantitative pair candidates: {summary['n_quantitative_pair_candidates']}")


if __name__ == "__main__":
    main()
