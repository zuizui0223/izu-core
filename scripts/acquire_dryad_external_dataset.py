#!/usr/bin/env python3
"""Acquire every source-native file from a public Dryad dataset.

The resolver is intentionally version-aware. A DOI's latest metadata resource may
not own downloadable files, so visible versions are inspected from newest to
oldest until the latest resource with files is found. File-level public routes
and Dryad's zip-assembly presigned URLs are both tried. Raw files remain workflow
artifacts; only compact inventories should be committed.
"""
from __future__ import annotations

import argparse
import csv
import hashlib
import io
import json
import math
import re
import urllib.error
import urllib.request
import zipfile
from pathlib import Path
from typing import Any, Iterable, Mapping


USER_AGENT = "izu-core-source-audit/1.0 (+https://github.com/zuizui0223/izu-core)"


def make_opener() -> urllib.request.OpenerDirector:
    return urllib.request.build_opener(urllib.request.HTTPRedirectHandler())


def request_bytes(
    opener: urllib.request.OpenerDirector,
    url: str,
    *,
    accept: str = "application/octet-stream, */*;q=0.8",
    referer: str | None = None,
) -> bytes:
    headers = {"User-Agent": USER_AGENT, "Accept": accept, "Cache-Control": "no-cache"}
    if referer:
        headers["Referer"] = referer
    request = urllib.request.Request(url, headers=headers)
    try:
        with opener.open(request, timeout=120) as response:
            return response.read()
    except urllib.error.HTTPError as error:
        try:
            body = error.read(400).decode("utf-8", errors="replace")
        except Exception:
            body = ""
        body = re.sub(r"\s+", " ", body).strip()
        raise RuntimeError(f"HTTP {error.code} {error.reason}; body={body!r}") from error


def request_json(opener: urllib.request.OpenerDirector, url: str) -> Any:
    payload = request_bytes(opener, url, accept="application/json")
    return json.loads(payload.decode("utf-8"))


def walk(value: object) -> Iterable[object]:
    yield value
    if isinstance(value, dict):
        for child in value.values():
            yield from walk(child)
    elif isinstance(value, list):
        for child in value:
            yield from walk(child)


def link_href(value: object) -> str | None:
    if isinstance(value, str):
        return value
    if isinstance(value, dict) and isinstance(value.get("href"), str):
        return str(value["href"])
    return None


def embedded_rows(value: object) -> list[dict[str, Any]]:
    if not isinstance(value, dict):
        return []
    embedded = value.get("_embedded")
    if not isinstance(embedded, dict):
        return []
    rows = []
    for candidate in embedded.values():
        if isinstance(candidate, list):
            rows.extend(row for row in candidate if isinstance(row, dict))
    return rows


def next_url(value: object) -> str | None:
    if not isinstance(value, dict):
        return None
    links = value.get("_links")
    if not isinstance(links, dict):
        return None
    return link_href(links.get("next"))


def paged_rows(opener: urllib.request.OpenerDirector, url: str) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    rows: list[dict[str, Any]] = []
    errors: list[dict[str, str]] = []
    seen: set[str] = set()
    current: str | None = url
    while current and current not in seen and len(seen) < 100:
        seen.add(current)
        try:
            page = request_json(opener, current)
        except Exception as error:
            errors.append({"url": current, "error": repr(error)})
            break
        rows.extend(embedded_rows(page))
        current = next_url(page)
    return rows, errors


def id_from_links(row: Mapping[str, object], kind: str) -> int | None:
    direct = row.get("id")
    if isinstance(direct, int):
        return direct
    expression = re.compile(rf"/{re.escape(kind)}/(\d+)(?:/|$)")
    for value in walk(row.get("_links")):
        if not isinstance(value, str):
            continue
        match = expression.search(value)
        if match:
            return int(match.group(1))
    return None


def numeric_version(row: Mapping[str, object]) -> float:
    for key in ("version", "versionNumber", "version_number"):
        value = row.get(key)
        try:
            number = float(value)
        except (TypeError, ValueError):
            continue
        if math.isfinite(number):
            return number
    return -math.inf


def version_sort_key(row: Mapping[str, object]) -> tuple[float, int]:
    return numeric_version(row), id_from_links(row, "versions") or -1


def safe_name(value: str) -> str:
    name = Path(str(value)).name
    name = re.sub(r"[^A-Za-z0-9._+() -]+", "_", name).strip(" ._")
    return name or "dryad_file"


def source_filename(row: Mapping[str, object]) -> str:
    for key in ("path", "name", "filename", "downloadFilename", "download_filename"):
        value = str(row.get(key) or "").strip()
        if value:
            return Path(value).name
    file_id = id_from_links(row, "files")
    return f"file_{file_id or 'unknown'}"


def all_hrefs(row: Mapping[str, object]) -> list[str]:
    urls = []
    for value in walk(row.get("_links")):
        if not isinstance(value, str):
            continue
        if "download" in value.casefold() or "/file_stream/" in value:
            if value not in urls:
                urls.append(value)
    return urls


def zip_info_urls(info: object, filename: str) -> list[str]:
    urls = []
    if not isinstance(info, list):
        return urls
    for row in info:
        if not isinstance(row, dict):
            continue
        candidate_name = Path(str(row.get("filename") or "")).name
        url = str(row.get("url") or "")
        if url and candidate_name == filename:
            urls.append(url)
    return urls


def looks_html(payload: bytes) -> bool:
    prefix = payload[:500].lstrip().casefold()
    return prefix.startswith(b"<!doctype html") or prefix.startswith(b"<html")


def valid_payload(filename: str, payload: bytes) -> tuple[bool, str]:
    if not payload:
        return False, "empty_response"
    if looks_html(payload):
        return False, "html_response"
    suffix = Path(filename).suffix.casefold()
    if suffix in {".xlsx", ".xlsm"}:
        try:
            with zipfile.ZipFile(io.BytesIO(payload)) as archive:
                names = archive.namelist()
        except zipfile.BadZipFile:
            return False, "invalid_xlsx_zip"
        if "[Content_Types].xml" not in names or not any(name.startswith("xl/") for name in names):
            return False, "missing_xlsx_structure"
    elif suffix == ".zip":
        if not zipfile.is_zipfile(io.BytesIO(payload)):
            return False, "invalid_zip"
    return True, "accepted"


def preview_delimited(path: Path) -> dict[str, object]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    delimiter = "\t" if path.suffix.casefold() == ".tsv" else ","
    try:
        delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",\t;").delimiter
    except csv.Error:
        pass
    preview = []
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    for index, row in enumerate(reader):
        preview.append(row[:30])
        if index >= 5:
            break
    return {"kind": "delimited", "delimiter": delimiter, "preview": preview}


def preview_xlsx(path: Path) -> dict[str, object]:
    try:
        import openpyxl
    except ImportError:
        return {"kind": "xlsx", "preview_error": "openpyxl_not_installed"}
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = []
    for worksheet in workbook.worksheets:
        preview = []
        for index, row in enumerate(worksheet.iter_rows(values_only=True)):
            preview.append(list(row[:30]))
            if index >= 5:
                break
        sheets.append(
            {
                "sheet": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "preview": preview,
            }
        )
    workbook.close()
    return {"kind": "xlsx", "sheets": sheets}


def safe_extract_zip(path: Path, destination: Path) -> list[str]:
    destination.mkdir(parents=True, exist_ok=True)
    members = []
    with zipfile.ZipFile(path) as archive:
        for member in archive.infolist():
            member_path = Path(member.filename)
            if member_path.is_absolute() or ".." in member_path.parts:
                continue
            members.append(member.filename)
            if member.is_dir():
                continue
            target = destination / member_path
            target.parent.mkdir(parents=True, exist_ok=True)
            target.write_bytes(archive.read(member))
    return members


def preview_file(path: Path, extraction_root: Path) -> dict[str, object]:
    suffix = path.suffix.casefold()
    if suffix in {".csv", ".tsv", ".txt"}:
        return preview_delimited(path)
    if suffix in {".xlsx", ".xlsm"}:
        return preview_xlsx(path)
    if suffix == ".zip" or zipfile.is_zipfile(path):
        return {"kind": "zip", "archive_members": safe_extract_zip(path, extraction_root / path.stem)}
    return {"kind": "other"}


def try_json(opener: urllib.request.OpenerDirector, url: str, errors: list[dict[str, str]], stage: str) -> object:
    try:
        return request_json(opener, url)
    except Exception as error:
        errors.append({"stage": stage, "url": url, "error": repr(error)})
        return {}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    args = parser.parse_args()

    config = json.loads(args.config.read_text(encoding="utf-8"))
    args.output_dir.mkdir(parents=True, exist_ok=True)
    raw_dir = args.output_dir / "files"
    extraction_root = args.output_dir / "extracted"
    raw_dir.mkdir(parents=True, exist_ok=True)
    opener = make_opener()
    errors: list[dict[str, str]] = []

    metadata = try_json(opener, str(config["api_dataset_url"]), errors, "dataset_metadata")
    version_rows, version_page_errors = paged_rows(opener, str(config["api_versions_url"]))
    errors.extend({"stage": "version_page", **row} for row in version_page_errors)
    versions = sorted(version_rows, key=version_sort_key, reverse=True)

    selected_version: dict[str, Any] | None = None
    selected_version_id: int | None = None
    selected_files: list[dict[str, Any]] = []
    file_listing_audit: list[dict[str, object]] = []

    for version in versions:
        version_id = id_from_links(version, "versions")
        if version_id is None:
            continue
        rows, page_errors = paged_rows(
            opener,
            f"https://datadryad.org/api/v2/versions/{version_id}/files?per_page=100",
        )
        file_listing_audit.append(
            {
                "version_id": version_id,
                "version_number": numeric_version(version),
                "n_file_rows": len(rows),
                "page_errors": page_errors,
            }
        )
        if rows:
            selected_version = version
            selected_version_id = version_id
            selected_files = rows
            break

    zip_info: object = {}
    if selected_version_id is not None:
        zip_info = try_json(
            opener,
            f"https://datadryad.org/downloads/zip_assembly_info/{selected_version_id}.json",
            errors,
            "zip_assembly_info",
        )

    inventory = []
    used_names: set[str] = set()
    for source_row in selected_files:
        filename = source_filename(source_row)
        local_name = safe_name(filename)
        stem = Path(local_name).stem
        suffix = Path(local_name).suffix
        counter = 2
        while local_name in used_names:
            local_name = f"{stem}_{counter}{suffix}"
            counter += 1
        used_names.add(local_name)
        file_id = id_from_links(source_row, "files")
        candidates = all_hrefs(source_row)
        if file_id is not None:
            for url in (
                f"https://datadryad.org/api/v2/files/{file_id}/download",
                f"https://datadryad.org/downloads/file_stream/{file_id}",
                f"https://datadryad.org/stash/downloads/file_stream/{file_id}",
            ):
                if url not in candidates:
                    candidates.append(url)
        for url in zip_info_urls(zip_info, filename):
            if url not in candidates:
                candidates.insert(0, url)

        payload: bytes | None = None
        successful_url: str | None = None
        attempts = []
        for url in candidates:
            try:
                candidate = request_bytes(opener, url, referer=str(config["landing_page_url"]))
                accepted, reason = valid_payload(filename, candidate)
                attempts.append({"url": url, "status": reason, "size": len(candidate)})
                if accepted:
                    payload = candidate
                    successful_url = url
                    break
            except Exception as error:
                attempts.append({"url": url, "status": "request_failed", "error": repr(error)})
        row: dict[str, object] = {
            "file_id": file_id,
            "source_filename": filename,
            "source_size": source_row.get("size"),
            "source_digest": source_row.get("digest"),
            "download_attempts": attempts,
        }
        if payload is None or successful_url is None:
            row["status"] = "download_failed"
            inventory.append(row)
            continue
        destination = raw_dir / local_name
        destination.write_bytes(payload)
        row.update(
            {
                "status": "downloaded",
                "local_name": local_name,
                "size_downloaded": len(payload),
                "sha256": hashlib.sha256(payload).hexdigest(),
                "successful_download_url": successful_url,
                "preview": preview_file(destination, extraction_root),
            }
        )
        inventory.append(row)

    summary = {
        "status": "acquired" if inventory and all(row.get("status") == "downloaded" for row in inventory) else "partial_or_failed",
        "source_id": config["source_id"],
        "dataset_doi": config["dataset_doi"],
        "metadata": metadata,
        "visible_versions": versions,
        "file_listing_audit": file_listing_audit,
        "selected_version_id": selected_version_id,
        "selected_version": selected_version,
        "n_source_files": len(selected_files),
        "n_downloaded": sum(row.get("status") == "downloaded" for row in inventory),
        "files": inventory,
        "errors": errors,
        "expected_source_components": config.get("expected_source_components"),
        "claim_boundary": config["claim_boundary"],
    }
    (args.output_dir / "source_inventory.json").write_text(
        json.dumps(summary, indent=2, ensure_ascii=False, default=str) + "\n",
        encoding="utf-8",
    )
    print(f"selected version: {selected_version_id}")
    print(f"files downloaded: {summary['n_downloaded']}/{summary['n_source_files']}")
    if not inventory or summary["n_downloaded"] != summary["n_source_files"]:
        raise RuntimeError("Dryad acquisition incomplete; see source_inventory.json")


if __name__ == "__main__":
    main()
