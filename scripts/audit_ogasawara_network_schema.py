#!/usr/bin/env python3
"""Audit downloaded Ogasawara tables without imposing an unverified schema."""
from __future__ import annotations

import argparse
import csv
import json
import re
from pathlib import Path
from typing import Iterable


ALIASES = {
    "island": ("island", "island_name"),
    "site": ("site", "plot", "locality", "location"),
    "season": ("season", "month", "survey_date", "date", "period"),
    "habitat": ("habitat", "forest", "disturbance", "disturbed", "vegetation"),
    "anole_context": ("anole", "lizard", "predation", "predator"),
    "plant": ("plant", "plant_species", "flower", "floral_species"),
    "pollinator": ("pollinator", "pollinator_species", "visitor", "insect", "animal"),
    "interaction_count": (
        "interaction_count",
        "legitimate_interaction",
        "visit_count",
        "visitation_frequency",
        "frequency",
        "count",
        "visits",
    ),
}


def norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().casefold()).strip("_")


def map_fields(headers: Iterable[object]) -> dict[str, list[str]]:
    normalized = {norm(value): str(value) for value in headers if str(value or "").strip()}
    matches: dict[str, list[str]] = {}
    for role, aliases in ALIASES.items():
        found = []
        for key, original in normalized.items():
            if any(alias == key or alias in key for alias in aliases):
                found.append(original)
        matches[role] = found
    return matches


def first_nonempty_row(rows: Iterable[Iterable[object]]) -> list[object]:
    for row in rows:
        values = list(row)
        if any(str(value or "").strip() for value in values):
            return values
    return []


def audit_delimited(path: Path) -> list[dict[str, object]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",\t;").delimiter
    except csv.Error:
        pass
    reader = csv.reader(text.splitlines(), delimiter=delimiter)
    rows = list(reader)
    headers = first_nonempty_row(rows)
    return [table_record(path, None, headers, max(0, len(rows) - 1), len(headers), delimiter)]


def audit_xlsx(path: Path) -> list[dict[str, object]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    records = []
    for worksheet in workbook.worksheets:
        headers = first_nonempty_row(worksheet.iter_rows(values_only=True))
        records.append(
            table_record(
                path,
                worksheet.title,
                headers,
                worksheet.max_row,
                worksheet.max_column,
                None,
            )
        )
    workbook.close()
    return records


def table_record(
    path: Path,
    sheet: str | None,
    headers: list[object],
    n_rows: int,
    n_columns: int,
    delimiter: str | None,
) -> dict[str, object]:
    roles = map_fields(headers)
    matched = {role for role, values in roles.items() if values}
    long_core = {"plant", "pollinator", "interaction_count"}.issubset(matched)
    context_core = {"island", "site", "season"}.issubset(matched)
    first_header = norm(headers[0]) if headers else ""
    wide_hint = n_columns >= 4 and (
        "plant" in first_header or "species" in first_header or first_header in {"", "taxon"}
    ) and not long_core
    return {
        "file": str(path),
        "sheet": sheet,
        "n_rows_reported": n_rows,
        "n_columns_reported": n_columns,
        "delimiter": delimiter,
        "headers": [str(value or "") for value in headers],
        "role_matches": roles,
        "long_network_candidate": long_core,
        "contextual_long_network_candidate": long_core and context_core,
        "wide_interaction_matrix_candidate": wide_hint,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", type=Path, default=Path("artifacts/ogasawara_zenodo"))
    parser.add_argument("--output", type=Path, default=Path("artifacts/ogasawara_zenodo/schema_audit.json"))
    args = parser.parse_args()

    tables = []
    paths = sorted(
        path
        for path in args.input_dir.rglob("*")
        if path.is_file()
        and path.name not in {"source_inventory.json", "schema_audit.json", "acquisition_failure.json"}
    )
    for path in paths:
        suffix = path.suffix.lower()
        try:
            if suffix in {".csv", ".tsv", ".txt"}:
                tables.extend(audit_delimited(path))
            elif suffix in {".xlsx", ".xlsm"}:
                tables.extend(audit_xlsx(path))
        except Exception as error:
            tables.append({"file": str(path), "status": "table_audit_failed", "error": repr(error)})

    summary = {
        "status": "schema_candidates_found" if tables else "no_tabular_files_found",
        "n_tabular_units": len(tables),
        "n_contextual_long_candidates": sum(bool(row.get("contextual_long_network_candidate")) for row in tables),
        "n_long_candidates": sum(bool(row.get("long_network_candidate")) for row in tables),
        "n_wide_candidates": sum(bool(row.get("wide_interaction_matrix_candidate")) for row in tables),
        "tables": tables,
        "next_gate": (
            "Select source-defined table(s), verify legitimate-interaction semantics and sampling effort, then implement a source-specific parser."
        ),
        "claim_boundary": (
            "Header matching is a schema-discovery aid, not biological classification. A visit or interaction count is not pollinator effectiveness, effective dependency or reproductive success."
        ),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"tabular units: {len(tables)}")
    print(f"long candidates: {summary['n_long_candidates']}")
    print(f"wide candidates: {summary['n_wide_candidates']}")


if __name__ == "__main__":
    main()
