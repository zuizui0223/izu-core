#!/usr/bin/env python3
"""Classify Galápagos Dryad tables without guessing biological orientation."""
from __future__ import annotations

import argparse
import csv
import json
import math
import re
from pathlib import Path
from typing import Iterable, Sequence


ROLE_ALIASES = {
    "island": ("island", "isla"),
    "plant": ("plant", "plant_species", "plant_sp", "flowering_plant"),
    "pollinator": ("pollinator", "visitor", "animal", "insect"),
    "interaction_weight": ("weight", "frequency", "visits", "visit_count", "interaction", "abundance"),
    "sampling_effort": ("effort", "hours", "observation_time", "sampling_days", "transect"),
    "area": ("area", "island_area"),
    "isolation": ("isolation", "distance", "dist_mainland", "nearest_island"),
    "age": ("age", "island_age", "geological_age"),
    "elevation": ("elevation", "altitude", "maximum_height"),
    "native_status": ("native", "alien", "introduced", "endemic", "origin", "status"),
}


def norm(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().casefold()).strip("_")


def matched_roles(headers: Sequence[object]) -> dict[str, list[str]]:
    normalized = [(norm(header), str(header or "")) for header in headers if str(header or "").strip()]
    output = {}
    for role, aliases in ROLE_ALIASES.items():
        output[role] = [
            original
            for key, original in normalized
            if any(key == alias or alias in key for alias in aliases)
        ]
    return output


def finite_fraction(values: Iterable[object]) -> float:
    total = 0
    finite = 0
    for value in values:
        if value is None or str(value).strip() == "":
            continue
        total += 1
        try:
            number = float(value)
        except (TypeError, ValueError):
            continue
        if math.isfinite(number):
            finite += 1
    return finite / total if total else 0.0


def first_data_block(rows: Iterable[Sequence[object]]) -> tuple[list[object], list[list[object]]]:
    iterator = iter(rows)
    headers = []
    for row in iterator:
        values = list(row)
        if any(str(value or "").strip() for value in values):
            headers = values
            break
    sample = []
    for row in iterator:
        values = list(row)
        if any(str(value or "").strip() for value in values):
            sample.append(values)
        if len(sample) >= 30:
            break
    return headers, sample


def classify_table(
    *,
    path: Path,
    sheet: str | None,
    headers: list[object],
    sample_rows: list[list[object]],
    n_rows: int,
    n_columns: int,
    delimiter: str | None,
) -> dict[str, object]:
    roles = matched_roles(headers)
    matched = {role for role, values in roles.items() if values}
    long_edge = {"plant", "pollinator", "interaction_weight"}.issubset(matched)
    island_covariates = "island" in matched and len(matched & {"area", "isolation", "age", "elevation"}) >= 2
    effort_table = "sampling_effort" in matched and ("island" in matched or "plant" in matched)

    first_header = norm(headers[0]) if headers else ""
    numeric_columns = []
    for column in range(1, max(len(headers), n_columns)):
        values = [row[column] if column < len(row) else None for row in sample_rows]
        if finite_fraction(values) >= 0.8 and any(str(value or "").strip() for value in values):
            numeric_columns.append(column)
    explicit_plant_axis = "plant" in first_header or "flower" in first_header
    explicit_pollinator_axis = "pollinator" in first_header or "visitor" in first_header
    wide_matrix = n_columns >= 4 and len(numeric_columns) >= 3 and not long_edge
    orientation = (
        "plants_by_pollinators"
        if explicit_plant_axis
        else "pollinators_by_plants"
        if explicit_pollinator_axis
        else "unresolved"
    )

    source_label = "__".join(part for part in (path.stem, sheet or "") if part)
    return {
        "file": str(path),
        "sheet": sheet,
        "source_label": source_label,
        "n_rows_reported": n_rows,
        "n_columns_reported": n_columns,
        "delimiter": delimiter,
        "headers": [str(value or "") for value in headers],
        "role_matches": roles,
        "long_edge_list_candidate": long_edge,
        "island_covariate_candidate": island_covariates,
        "sampling_effort_candidate": effort_table,
        "wide_numeric_matrix_candidate": wide_matrix,
        "matrix_orientation": orientation,
        "numeric_sample_column_indices": numeric_columns,
        "analysis_admissible_matrix": wide_matrix and orientation != "unresolved",
    }


def read_delimited(path: Path) -> list[dict[str, object]]:
    text = path.read_text(encoding="utf-8-sig", errors="replace")
    delimiter = "\t" if path.suffix.casefold() == ".tsv" else ","
    try:
        delimiter = csv.Sniffer().sniff(text[:8192], delimiters=",\t;").delimiter
    except csv.Error:
        pass
    rows = list(csv.reader(text.splitlines(), delimiter=delimiter))
    headers, sample = first_data_block(rows)
    return [
        classify_table(
            path=path,
            sheet=None,
            headers=headers,
            sample_rows=sample,
            n_rows=len(rows),
            n_columns=max((len(row) for row in rows), default=0),
            delimiter=delimiter,
        )
    ]


def read_xlsx(path: Path) -> list[dict[str, object]]:
    import openpyxl

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    output = []
    for worksheet in workbook.worksheets:
        headers, sample = first_data_block(worksheet.iter_rows(values_only=True))
        output.append(
            classify_table(
                path=path,
                sheet=worksheet.title,
                headers=headers,
                sample_rows=sample,
                n_rows=worksheet.max_row,
                n_columns=worksheet.max_column,
                delimiter=None,
            )
        )
    workbook.close()
    return output


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input-dir", type=Path, default=Path("artifacts/galapagos_dryad"))
    parser.add_argument("--output", type=Path, default=Path("artifacts/galapagos_dryad/schema_audit.json"))
    args = parser.parse_args()

    tables = []
    errors = []
    for path in sorted(args.input_dir.rglob("*")):
        if not path.is_file() or path.name in {"source_inventory.json", "schema_audit.json"}:
            continue
        try:
            if path.suffix.casefold() in {".csv", ".tsv", ".txt"}:
                tables.extend(read_delimited(path))
            elif path.suffix.casefold() in {".xlsx", ".xlsm"}:
                tables.extend(read_xlsx(path))
        except Exception as error:
            errors.append({"file": str(path), "error": repr(error)})

    summary = {
        "status": "tabular_schema_audited" if tables else "no_tabular_sources_resolved",
        "n_tabular_units": len(tables),
        "n_long_edge_candidates": sum(bool(row["long_edge_list_candidate"]) for row in tables),
        "n_wide_matrix_candidates": sum(bool(row["wide_numeric_matrix_candidate"]) for row in tables),
        "n_admissible_oriented_matrices": sum(bool(row["analysis_admissible_matrix"]) for row in tables),
        "n_island_covariate_candidates": sum(bool(row["island_covariate_candidate"]) for row in tables),
        "n_sampling_effort_candidates": sum(bool(row["sampling_effort_candidate"]) for row in tables),
        "tables": tables,
        "errors": errors,
        "next_gate": (
            "Resolve source-defined island labels and sampling effort, then analyse only explicitly oriented matrices or long edge lists. Unresolved wide matrices remain blocked."
        ),
        "claim_boundary": (
            "Schema classification is not a biological result. Network matrices contain interactions or visits, not per-visit effectiveness, effective dependency or reproductive success. Island area, age and isolation are descriptive moderators unless the source design identifies them causally."
        ),
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(summary, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(f"tabular units: {summary['n_tabular_units']}")
    print(f"network candidates: {summary['n_long_edge_candidates'] + summary['n_wide_matrix_candidates']}")
    print(f"island covariate candidates: {summary['n_island_covariate_candidates']}")


if __name__ == "__main__":
    main()
