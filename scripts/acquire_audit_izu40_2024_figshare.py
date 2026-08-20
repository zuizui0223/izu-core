from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import re
import urllib.error
import urllib.request
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
DESIGN = ROOT / "data/design/abm_v8_izu40_figshare_source_gate_v1.json"
OUT = ROOT / "data/results/izu40_2024_figshare_source_audit.json"
RAW_DIR = ROOT / "data/external/izu40_2024"
API = "https://api.figshare.com/v2/articles/25025000"
USER_AGENT = "izu-core-source-audit/1.0"


def fetch_bytes(url: str) -> tuple[int | None, bytes | None, str | None]:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    try:
        with urllib.request.urlopen(request, timeout=90) as response:
            return int(response.status), response.read(), None
    except urllib.error.HTTPError as exc:
        return int(exc.code), None, str(exc)
    except Exception as exc:
        return None, None, f"{type(exc).__name__}: {exc}"


def normalize(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(value).lower()).strip("_")


def role_candidates(headers: list[str]) -> dict[str, list[str]]:
    normalized = [normalize(value) for value in headers]
    rules = {
        "network_or_site": ("network", "site", "island", "location", "local"),
        "time": ("season", "survey", "date", "month", "year", "time", "period"),
        "plant": ("plant", "flower"),
        "pollinator": ("pollinator", "visitor", "insect", "bee", "animal"),
        "weight": ("visit", "interaction", "frequency", "freq", "count", "abundance", "number"),
    }
    return {
        role: [header for header, key in zip(headers, normalized) if any(token in key for token in tokens)]
        for role, tokens in rules.items()
    }


def candidate_header(rows: list[list[object]]) -> tuple[int | None, list[str]]:
    for index, row in enumerate(rows, start=1):
        values = ["" if value is None else str(value).strip() for value in row]
        nonempty = [value for value in values if value]
        textual = [value for value in nonempty if not _is_number(value)]
        if len(nonempty) >= 2 and len(textual) >= min(2, len(nonempty)):
            return index, values
    return None, []


def _is_number(value: str) -> bool:
    try:
        number = float(value)
    except ValueError:
        return False
    return math.isfinite(number)


def inventory_xlsx(payload: bytes) -> dict:
    book = load_workbook(io.BytesIO(payload), read_only=True, data_only=False)
    sheets = []
    for sheet in book.worksheets:
        preview_rows: list[list[object]] = []
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 8), values_only=True):
            preview_rows.append(list(row))
        header_row, headers = candidate_header(preview_rows)
        sheets.append({
            "name": sheet.title,
            "max_row": sheet.max_row,
            "max_column": sheet.max_column,
            "candidate_header_row": header_row,
            "candidate_headers": headers,
            "field_role_candidates": role_candidates(headers),
        })
    book.close()
    return {"format": "xlsx", "sheet_count": len(sheets), "sheets": sheets, "target_metrics_calculated": False}


def decode_text(payload: bytes) -> tuple[str, str]:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return payload.decode(encoding), encoding
        except UnicodeDecodeError:
            continue
    raise RuntimeError("text source could not be decoded")


def inventory_delimited(payload: bytes) -> dict:
    text, encoding = decode_text(payload)
    lines = text.splitlines()
    sample = "\n".join(lines[:30])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="\t,;")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = "\t" if "\t" in (lines[0] if lines else "") else ","
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    header_row, headers = candidate_header([list(row) for row in rows[:8]])
    return {
        "format": "delimited_text",
        "encoding": encoding,
        "delimiter_repr": repr(delimiter),
        "row_count": len(rows),
        "column_count_first_row": len(rows[0]) if rows else 0,
        "candidate_header_row": header_row,
        "candidate_headers": headers,
        "field_role_candidates": role_candidates(headers),
        "target_metrics_calculated": False,
    }


def inventory_code(payload: bytes) -> dict:
    text, encoding = decode_text(payload)
    indicators = []
    tokens = ("network", "interaction", "site", "island", "season", "survey", "pollinator", "plant", "bipartite")
    for index, line in enumerate(text.splitlines(), start=1):
        lower = line.lower()
        if any(token in lower for token in tokens):
            indicators.append({"line": index, "text": line[:300]})
    return {
        "format": "source_code",
        "encoding": encoding,
        "line_count": len(text.splitlines()),
        "structural_indicator_lines": indicators[:160],
        "target_metrics_calculated": False,
    }


def structured_roles_visible(inventory: dict) -> bool:
    candidates = []
    if inventory.get("format") == "xlsx":
        candidates = [sheet.get("field_role_candidates", {}) for sheet in inventory.get("sheets", [])]
    elif inventory.get("format") == "delimited_text":
        candidates = [inventory.get("field_role_candidates", {})]
    for roles in candidates:
        if roles.get("plant") and roles.get("pollinator") and (roles.get("network_or_site") or roles.get("time")):
            return True
    return False


def write(payload: dict) -> None:
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
    print(json.dumps(payload, indent=2, ensure_ascii=False))


def main() -> None:
    design = json.loads(DESIGN.read_text())
    status, metadata_bytes, metadata_error = fetch_bytes(API)
    if status != 200 or metadata_bytes is None:
        write({
            "schema_version": "1.0",
            "analysis": "izu40_2024_figshare_source_audit",
            "status": "blocked_figshare_metadata_not_recovered",
            "metadata_http_status": status,
            "metadata_error": metadata_error,
            "source_admission_succeeds": False,
            "target_metrics_calculated": False,
        })
        return

    metadata = json.loads(metadata_bytes)
    files = metadata.get("files") or []
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    records = []
    blocked = []
    structured_visible = False
    code_structure_visible = False

    for row in files:
        name = str(row.get("name") or row.get("key") or "").strip()
        download_url = row.get("download_url") or (row.get("links") or {}).get("download") or (row.get("links") or {}).get("content")
        record = {
            "figshare_file_id": row.get("id"),
            "name": name,
            "metadata_bytes": row.get("size"),
            "metadata_supplied_md5": row.get("supplied_md5") or row.get("computed_md5"),
            "download_url": download_url,
        }
        if not name or not download_url:
            record["error"] = "missing filename or download URL in Figshare metadata"
            blocked.append(name or "<unnamed>")
            records.append(record)
            continue
        file_status, payload, error = fetch_bytes(download_url)
        record["http_status"] = file_status
        record["error"] = error
        if file_status != 200 or payload is None:
            blocked.append(name)
            records.append(record)
            continue
        actual_md5 = hashlib.md5(payload).hexdigest()
        actual_sha256 = hashlib.sha256(payload).hexdigest()
        supplied = str(record["metadata_supplied_md5"] or "")
        checksum_match = (not supplied) or supplied == actual_md5
        record.update({
            "bytes": len(payload),
            "md5": actual_md5,
            "sha256": actual_sha256,
            "metadata_md5_match": checksum_match,
        })
        if row.get("size") is not None and int(row["size"]) != len(payload):
            record["size_match"] = False
            blocked.append(name)
        else:
            record["size_match"] = True
        if not checksum_match:
            blocked.append(name)
        (RAW_DIR / name).write_bytes(payload)

        suffix = Path(name).suffix.lower()
        try:
            if suffix == ".xlsx":
                inventory = inventory_xlsx(payload)
                record["schema_inventory"] = inventory
                structured_visible = structured_visible or structured_roles_visible(inventory)
            elif suffix in {".csv", ".tsv", ".txt"}:
                inventory = inventory_delimited(payload)
                record["schema_inventory"] = inventory
                structured_visible = structured_visible or structured_roles_visible(inventory)
            elif suffix in {".r", ".rmd", ".py"}:
                inventory = inventory_code(payload)
                record["code_inventory"] = inventory
                code_structure_visible = code_structure_visible or bool(inventory["structural_indicator_lines"])
            else:
                record["inventory"] = {
                    "format": suffix.lstrip(".") or "unknown",
                    "structure_not_interpreted_in_source_gate": True,
                    "target_metrics_calculated": False,
                }
        except Exception as exc:
            record["inventory_error"] = f"{type(exc).__name__}: {exc}"
        records.append(record)

    source_bytes_ok = bool(files) and not blocked and len(records) == len(files)
    reconstruction_structure_visible = structured_visible or code_structure_visible
    admission = source_bytes_ok and reconstruction_structure_visible
    write({
        "schema_version": "1.0",
        "analysis": "izu40_2024_figshare_source_audit",
        "status": (
            "source_admitted_figshare_data_code_structure_before_v8_targets"
            if admission else "blocked_izu40_figshare_source_or_reconstruction_structure_incomplete"
        ),
        "figshare_article_id": design["candidate_system"]["figshare_article_id"],
        "figshare_doi": design["candidate_system"]["figshare_doi"],
        "metadata_title": metadata.get("title"),
        "metadata_doi": metadata.get("doi"),
        "metadata_version": metadata.get("version"),
        "metadata_bytes": len(metadata_bytes),
        "metadata_sha256": hashlib.sha256(metadata_bytes).hexdigest(),
        "file_count": len(files),
        "blocked_files": sorted(set(blocked)),
        "files": records,
        "source_bytes_ok": source_bytes_ok,
        "structured_interaction_roles_visible": structured_visible,
        "source_code_structure_visible": code_structure_visible,
        "reconstruction_structure_visible": reconstruction_structure_visible,
        "source_admission_succeeds": admission,
        "target_metrics_calculated": False,
        "published_scope": design["candidate_system"]["published_scope"],
        "independence_boundary": design["independence_boundary"],
        "claim_boundary": "Source bytes/schema/code inventory only. No network metric, support estimand, empirical range, or ABM v8 predictive fit was calculated.",
    })


if __name__ == "__main__":
    main()
