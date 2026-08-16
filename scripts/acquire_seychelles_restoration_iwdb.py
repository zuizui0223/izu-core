#!/usr/bin/env python3
"""Acquire Kaiser-Bunbury et al. 2017 Seychelles restoration-network bytes.

Primary authority is the Interaction Web DataBase (IWDB). If that host is
unreachable, use the configured Zenodo repository snapshot strictly as a byte
transport fallback and preserve the distinction in the source inventory.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from html.parser import HTMLParser
from pathlib import Path


class LinkParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__(); self.links=[]; self._href=None; self._text=[]
    def handle_starttag(self, tag, attrs):
        if tag.lower()=="a": self._href=dict(attrs).get("href"); self._text=[]
    def handle_data(self, data):
        if self._href is not None: self._text.append(data)
    def handle_endtag(self, tag):
        if tag.lower()=="a" and self._href is not None:
            self.links.append((self._href," ".join(self._text).strip())); self._href=None; self._text=[]


def request_bytes(url: str, timeout: int = 30) -> bytes:
    req=urllib.request.Request(url,headers={"User-Agent":"izu-core-source-audit/1.0 (+https://github.com/zuizui0223/izu-core)","Accept":"*/*"})
    with urllib.request.urlopen(req,timeout=timeout) as response: return response.read()


def safe_name(url: str, fallback: str) -> str:
    name=Path(urllib.parse.unquote(urllib.parse.urlparse(url).path)).name or fallback
    return re.sub(r"[^A-Za-z0-9._-]+","_",name)


def workbook_inventory(path: Path) -> dict[str,object]:
    suffix=path.suffix.lower()
    if suffix==".xlsx":
        import openpyxl
        wb=openpyxl.load_workbook(path,read_only=True,data_only=True); sheets=[]
        for ws in wb.worksheets:
            preview=[]
            for i,row in enumerate(ws.iter_rows(values_only=True)):
                preview.append(list(row))
                if i>=2: break
            sheets.append({"sheet":ws.title,"max_row":ws.max_row,"max_column":ws.max_column,"preview":preview})
        wb.close(); return {"format":"xlsx","sheets":sheets}
    if suffix==".xls":
        import xlrd
        book=xlrd.open_workbook(path); sheets=[]
        for sheet in book.sheets():
            sheets.append({"sheet":sheet.name,"max_row":sheet.nrows,"max_column":sheet.ncols,"preview":[sheet.row_values(i) for i in range(min(sheet.nrows,3))]})
        return {"format":"xls","sheets":sheets}
    return {"format":suffix.lstrip(".") or "unknown"}


def primary_attempt(page_url: str, raw_dir: Path) -> dict[str,object]:
    state={"route":"primary_iwdb","status":"not_attempted","error":None,"page_sha256":None,"candidate_links":[],"downloaded_spreadsheets":[]}
    try:
        page=request_bytes(page_url,timeout=30); state["status"]="page_recovered"; state["page_sha256"]=hashlib.sha256(page).hexdigest()
        parser=LinkParser(); parser.feed(page.decode("utf-8",errors="replace")); links=[]
        for href,label in parser.links:
            absolute=urllib.parse.urljoin(page_url,href); combined=f"{href} {label}".lower()
            if any(token in combined for token in (".xlsx",".xls","64 network","visits","visitfreq")):
                links.append({"href":href,"label":label,"url":absolute})
        state["candidate_links"]=links
        sheets=[row for row in links if re.search(r"\.xlsx?(?:$|[?#])",row["url"],re.I)]
        for i,row in enumerate(sheets,1):
            payload=request_bytes(row["url"],timeout=60); name=safe_name(row["url"],f"iwdb_{i}.xlsx"); dest=raw_dir/name; dest.write_bytes(payload)
            if len(payload)<1024: raise ValueError(f"implausibly small spreadsheet {name}")
            state["downloaded_spreadsheets"].append({"name":name,"label":row["label"],"url":row["url"],"bytes":len(payload),"sha256":hashlib.sha256(payload).hexdigest(),"workbook":workbook_inventory(dest)})
        if state["downloaded_spreadsheets"]: state["status"]="source_spreadsheets_recovered"
        elif state["status"]=="page_recovered": state["status"]="page_recovered_no_spreadsheet_href"
    except Exception as exc:
        state["status"]="transport_failed"; state["error"]=f"{type(exc).__name__}: {exc}"
    return state


def fallback_attempt(config: dict[str,object], outdir: Path) -> dict[str,object]:
    fb=dict(config["transport_fallback"]); api=str(fb["api_url"]); state={"route":"zenodo_transport_fallback","status":"not_attempted","record_doi":fb["doi"],"record_id":fb["record_id"],"files":[],"relevant_archive_members":[]}
    metadata=json.loads(request_bytes(api,timeout=60).decode("utf-8")); files=metadata.get("files") or []
    if not files: raise ValueError("Zenodo fallback exposes no files")
    fallback_dir=outdir/"fallback"; fallback_dir.mkdir(parents=True,exist_ok=True)
    for item in files:
        links=item.get("links") or {}; url=links.get("self") or links.get("download") or item.get("links",{}).get("content")
        if not url: continue
        payload=request_bytes(str(url),timeout=120); name=str(item.get("key") or item.get("filename") or f"zenodo_{item.get('id')}"); dest=fallback_dir/safe_name(name,"zenodo_file"); dest.write_bytes(payload)
        row={"name":dest.name,"url":url,"bytes":len(payload),"sha256":hashlib.sha256(payload).hexdigest(),"zenodo_checksum":item.get("checksum")}
        if zipfile.is_zipfile(dest):
            extract=fallback_dir/(dest.stem+"_extracted"); extract.mkdir(exist_ok=True)
            with zipfile.ZipFile(dest) as archive:
                members=[]
                for info in archive.infolist():
                    if info.is_dir(): continue
                    member={"name":info.filename,"bytes":info.file_size}
                    low=info.filename.lower()
                    if any(token in low for token in ("seychell","kaiser","restor","network")):
                        state["relevant_archive_members"].append(member)
                    members.append(member)
                archive.extractall(extract); row["archive_members_count"]=len(members)
        state["files"].append(row)
    state["status"]="transport_recovered"
    state["provenance_boundary"]=fb["scope"]
    return state


def main() -> None:
    ap=argparse.ArgumentParser(description=__doc__); ap.add_argument("--config",type=Path,default=Path("config/seychelles_restoration_network_iwdb_source.json")); ap.add_argument("--outdir",type=Path,default=Path("artifacts/seychelles_restoration_iwdb")); args=ap.parse_args()
    config=json.loads(args.config.read_text(encoding="utf-8")); args.outdir.mkdir(parents=True,exist_ok=True); raw=args.outdir/"files"; raw.mkdir(parents=True,exist_ok=True)
    primary=primary_attempt(str(config["source_page"]),raw)
    fallback=None
    if primary["status"]!="source_spreadsheets_recovered": fallback=fallback_attempt(config,args.outdir)
    summary={"schema_version":"1.1","source_id":config["source_id"],"article_doi":config["article_doi"],"source_authority":{"database":config["source_database"],"page":config["source_page"]},"primary_attempt":primary,"fallback_attempt":fallback,"expected_scale":{k:config[k] for k in ("expected_networks","expected_sites","expected_months","expected_observation_hours","expected_unique_links","expected_pollinator_visits")},"claim_boundary":config["claim_boundary"]}
    (args.outdir/"source_inventory.json").write_text(json.dumps(summary,indent=2,ensure_ascii=False,default=str)+"\n",encoding="utf-8")
    print(json.dumps({"primary":primary,"fallback":fallback},indent=2,ensure_ascii=False,default=str))

if __name__=="__main__": main()
