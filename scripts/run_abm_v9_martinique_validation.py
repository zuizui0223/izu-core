from __future__ import annotations

import io
import json
import math
import random
import statistics
import urllib.request
import hashlib
import importlib.util
import sys
from itertools import combinations
from pathlib import Path

from openpyxl import load_workbook

from channel_id.external_archipelago_network import WeightedNetwork

ROOT = Path(__file__).resolve().parents[1]
DESIGN = ROOT / "data/design/abm_v9_martinique_validation_v1.json"
RECONSTRUCTION = ROOT / "data/design/abm_v9_martinique_2025_reconstruction_v1.json"
V9_SCRIPT = ROOT / "scripts/run_constraint_mechanism_abm_v9_local_plant_opportunity.py"
V8_SCRIPT = ROOT / "scripts/run_constraint_mechanism_abm_v8_pair_support.py"
V6_SCRIPT = ROOT / "scripts/run_constraint_mechanism_abm_v6_local_support.py"
V5_SCRIPT = ROOT / "scripts/run_constraint_mechanism_abm_v5_hierarchical_context.py"
OUT = ROOT / "data/results/abm_v9_martinique_validation.json"
INTERACTION_URL = "https://search-data.ubfc.fr/dl_data.php?file=597"
SAMPLING_URL = "https://search-data.ubfc.fr/dl_data.php?file=601"
INTERACTION_SHA = "9a001287bf64d51cbdefee1164579398e0cf5053efbfc04d1f8bcf9338626753"
SAMPLING_SHA = "e3f82dc81749d7c759dbb62fc2e40ceeff9382758a3114c63c57553d15c2327d"
USER_AGENT = "izu-core-martinique-v9-validation/1.0"
MISSING = {"", "na", "n/a", "nan", "none", "null"}
SEED = 20260821


def load_module(path: Path, name: str):
    spec = importlib.util.spec_from_file_location(name, path)
    module = importlib.util.module_from_spec(spec)
    sys.modules[spec.name] = module
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def fetch(url: str, expected_sha: str) -> bytes:
    request = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(request, timeout=120) as response:
        payload = response.read()
    actual = hashlib.sha256(payload).hexdigest()
    if actual != expected_sha:
        raise RuntimeError(f"Martinique source checksum drift: {url} {actual}")
    return payload


def rows_from_sheet(payload: bytes, sheet_name: str) -> list[dict[str, object]]:
    book = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
    if sheet_name not in book.sheetnames:
        raise RuntimeError(f"missing source sheet: {sheet_name}")
    sheet = book[sheet_name]
    iterator = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(iterator)]
    rows: list[dict[str, object]] = []
    for raw in iterator:
        if not any(value not in (None, "") for value in raw):
            continue
        rows.append({header: value for header, value in zip(headers, raw) if header})
    book.close()
    return rows


def clean(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().split())


def identity(value: object) -> str:
    text = clean(value)
    return "" if text.casefold() in MISSING else text


def numeric(value: object) -> float | None:
    text = clean(value)
    if text.casefold() in MISSING:
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def shannon(weights) -> float:
    positive = [float(value) for value in weights if float(value) > 0.0]
    total = sum(positive)
    if total <= 0.0:
        return 0.0
    return -sum((value / total) * math.log(value / total) for value in positive)


def percentile(values: list[float], q: float) -> float:
    if not values:
        raise ValueError("percentile requires at least one value")
    ordered = sorted(float(value) for value in values)
    if len(ordered) == 1:
        return ordered[0]
    position = (len(ordered) - 1) * q
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return ordered[lower]
    fraction = position - lower
    return ordered[lower] * (1.0 - fraction) + ordered[upper] * fraction


def envelope(values: list[float]) -> dict[str, float]:
    return {
        "p2.5": percentile(values, 0.025),
        "median": percentile(values, 0.5),
        "p97.5": percentile(values, 0.975),
    }


def inside(value: float, interval: dict[str, float]) -> bool:
    return interval["p2.5"] <= float(value) <= interval["p97.5"]


def pair_set(network: WeightedNetwork | None) -> set[tuple[str, str]]:
    if network is None:
        return set()
    return {
        (plant, pollinator)
        for i, plant in enumerate(network.plant_names)
        for j, pollinator in enumerate(network.pollinator_names)
        if network.matrix[i][j] > 0.0
    }


def weighted_shannon(network: WeightedNetwork | None) -> float:
    if network is None:
        return 0.0
    return shannon(value for row in network.matrix for value in row)


def jaccard_turnover(pair_sets: list[set[tuple[str, str]]]) -> float:
    values = []
    for left, right in combinations(pair_sets, 2):
        union = left | right
        values.append(0.0 if not union else 1.0 - len(left & right) / len(union))
    return float(statistics.mean(values)) if values else 0.0


def build_empirical(design: dict, reconstruction: dict) -> tuple[dict, WeightedNetwork | None, dict]:
    interaction_payload = fetch(INTERACTION_URL, INTERACTION_SHA)
    sampling_payload = fetch(SAMPLING_URL, SAMPLING_SHA)
    interaction_rows = rows_from_sheet(interaction_payload, "Insects-Plants")
    floral_rows = rows_from_sheet(sampling_payload, "Floral_abundance")

    sites = list(reconstruction["network_unit"]["site_ids"])
    periods = list(reconstruction["network_unit"]["period_ids"])
    contexts = [(site, period) for site in sites for period in periods]
    context_set = set(contexts)
    if len(contexts) != 120:
        raise RuntimeError("frozen Martinique context grid is not 120")

    event_weights: dict[tuple[str, str], dict[tuple[str, str], float]] = {
        context: {} for context in contexts
    }
    blank_both = 0
    one_sided = 0
    for row in interaction_rows:
        context = (clean(row.get("Site")), clean(row.get("Period")))
        if context not in context_set:
            raise RuntimeError(f"interaction row outside frozen context grid: {context}")
        plant = identity(row.get("Plant_Best_ID"))
        insect = identity(row.get("Insect_Best_ID"))
        if not plant and not insect:
            blank_both += 1
            continue
        if not plant or not insect:
            one_sided += 1
            continue
        pair = (plant, insect)
        event_weights[context][pair] = event_weights[context].get(pair, 0.0) + 1.0
    if one_sided:
        raise RuntimeError(f"frozen one-sided identity invariant failed: {one_sided}")
    if blank_both != 821:
        raise RuntimeError(f"blank-both interaction row count drifted: {blank_both}")

    floral_plants: dict[tuple[str, str], set[str]] = {context: set() for context in contexts}
    floral_amounts: dict[tuple[str, str], dict[str, float]] = {context: {} for context in contexts}
    floral_na_placeholders = 0
    for row in floral_rows:
        context = (clean(row.get("Site")), clean(row.get("Period")))
        if context not in context_set:
            raise RuntimeError(f"floral row outside frozen context grid: {context}")
        plant = identity(row.get("Plant_Best_ID"))
        amount = numeric(row.get("Nb_Floral_unit"))
        if not plant:
            if amount is None:
                floral_na_placeholders += 1
                continue
            raise RuntimeError("floral row has numeric resource amount but missing plant identity")
        if amount is None or amount <= 0.0:
            raise RuntimeError("identified floral row lacks positive numeric Nb_Floral_unit")
        floral_plants[context].add(plant)
        floral_amounts[context][plant] = floral_amounts[context].get(plant, 0.0) + amount
    if floral_na_placeholders != 14:
        raise RuntimeError(f"floral NA placeholder count drifted: {floral_na_placeholders}")

    pooled_weights: dict[tuple[str, str], float] = {}
    for weights in event_weights.values():
        for pair, amount in weights.items():
            pooled_weights[pair] = pooled_weights.get(pair, 0.0) + amount
    if not pooled_weights:
        raise RuntimeError("Martinique pooled interaction baseline is empty")
    pooled_pairs = set(pooled_weights)
    pooled_plants = sorted({plant for plant, _ in pooled_pairs})
    pooled_pollinators = sorted({pollinator for _, pollinator in pooled_pairs})

    mismatch_rows = []
    for context in contexts:
        observed_plants = {plant for plant, _ in event_weights[context]}
        missing = sorted(observed_plants - floral_plants[context])
        if missing:
            mismatch_rows.append({
                "site": context[0],
                "period": context[1],
                "interaction_plants_not_in_independent_floral_opportunity": missing,
                "count": len(missing),
            })

    structural = {
        "passes": not mismatch_rows,
        "decision_if_failed": design["pre_target_structural_gate"]["failure_decision"],
        "mismatch_context_count": len(mismatch_rows),
        "mismatch_plant_endpoint_count": sum(row["count"] for row in mismatch_rows),
        "mismatches": mismatch_rows,
        "blank_both_sampling_rows": blank_both,
        "floral_na_placeholder_rows": floral_na_placeholders,
    }
    if mismatch_rows:
        return {
            "context_count": len(contexts),
            "pooled_positive_pair_count": len(pooled_pairs),
            "pooled_interacting_plant_count": len(pooled_plants),
            "pooled_interacting_insect_count": len(pooled_pollinators),
            "structural_gate": structural,
            "primary_estimands": None,
        }, None, {
            "event_weights": event_weights,
            "floral_plants": floral_plants,
            "floral_amounts": floral_amounts,
            "contexts": contexts,
        }

    plant_index = {name: index for index, name in enumerate(pooled_plants)}
    poll_index = {name: index for index, name in enumerate(pooled_pollinators)}
    matrix = [[0.0 for _ in pooled_pollinators] for _ in pooled_plants]
    for (plant, insect), amount in pooled_weights.items():
        matrix[plant_index[plant]][poll_index[insect]] = amount
    baseline = WeightedNetwork.from_rows(pooled_plants, pooled_pollinators, matrix)
    pooled_shannon = weighted_shannon(baseline)
    if pooled_shannon <= 0.0:
        raise RuntimeError("pooled Martinique Shannon is not positive")

    plant_fractions = []
    conditional_pair_fractions = []
    shannons = []
    pair_sets = []
    pollinator_fractions = []
    pair_counts = []
    floral_totals = []
    zero_interaction_contexts = 0
    pooled_plant_set = set(pooled_plants)
    pooled_pollinator_set = set(pooled_pollinators)

    for context in contexts:
        active_plants = floral_plants[context] & pooled_plant_set
        plant_fractions.append(len(active_plants) / len(pooled_plants))
        denominator_pairs = {pair for pair in pooled_pairs if pair[0] in active_plants}
        observed_pairs = set(event_weights[context])
        if denominator_pairs:
            conditional_pair_fractions.append(len(observed_pairs) / len(denominator_pairs))
        weights = event_weights[context]
        shannons.append(shannon(weights.values()))
        pair_sets.append(observed_pairs)
        pair_counts.append(len(observed_pairs))
        pollinators = {pollinator for _, pollinator in observed_pairs}
        pollinator_fractions.append(len(pollinators) / len(pooled_pollinators))
        floral_totals.append(sum(floral_amounts[context].values()))
        zero_interaction_contexts += not observed_pairs

    if not conditional_pair_fractions:
        raise RuntimeError("no empirical context has a defined conditional pair-support denominator")

    primary = {
        "median_active_interacting_plant_fraction": float(statistics.median(plant_fractions)),
        "median_pair_support_fraction_given_active_plants": float(statistics.median(conditional_pair_fractions)),
        "interaction_shannon_relative_context_range": (max(shannons) - min(shannons)) / pooled_shannon,
    }
    secondary = {
        "mean_pairwise_jaccard_turnover": jaccard_turnover(pair_sets),
        "active_pollinator_fraction_median": float(statistics.median(pollinator_fractions)),
        "active_pollinator_fraction_range": max(pollinator_fractions) - min(pollinator_fractions),
        "positive_pair_count_median": float(statistics.median(pair_counts)),
        "positive_pair_count_range": max(pair_counts) - min(pair_counts),
        "zero_interaction_context_count": zero_interaction_contexts,
        "floral_unit_total_median": float(statistics.median(floral_totals)),
        "floral_unit_total_range": max(floral_totals) - min(floral_totals),
        "defined_conditional_pair_context_count": len(conditional_pair_fractions),
        "shannon_min": min(shannons),
        "shannon_max": max(shannons),
    }
    empirical = {
        "context_count": len(contexts),
        "pooled_opportunity": {
            "positive_pair_count": len(pooled_pairs),
            "positive_plant_count": len(pooled_plants),
            "positive_insect_count": len(pooled_pollinators),
            "total_event_count": sum(pooled_weights.values()),
            "interaction_shannon": pooled_shannon,
        },
        "structural_gate": structural,
        "primary_estimands": primary,
        "secondary": secondary,
    }
    return empirical, baseline, {
        "event_weights": event_weights,
        "floral_plants": floral_plants,
        "floral_amounts": floral_amounts,
        "contexts": contexts,
    }


def fast_supported_context(v9, v8, v6, baseline: WeightedNetwork, *, support_seed: int, support_strength: float):
    active_plants = v9.draw_local_plant_indices(
        baseline,
        plant_seed=support_seed + v9.PLANT_SEED_OFFSET,
        support_strength=support_strength,
    )
    active_plant_set = set(active_plants)
    globally_active = v6.active_pollinator_indices(
        len(baseline.pollinator_names),
        rng=random.Random(support_seed),
        support_strength=support_strength,
    )
    active_pollinator_set = set(globally_active)
    pair_rng = random.Random(support_seed + v8.PAIR_SEED_OFFSET)
    keep_probability = 1.0 - support_strength
    combined_mask = []
    for row_index, row in enumerate(baseline.matrix):
        mask_row = []
        for column, value in enumerate(row):
            pair_active = False
            if column in active_pollinator_set and value > 0.0:
                pair_active = pair_rng.random() < keep_probability
            mask_row.append(pair_active and row_index in active_plant_set)
        combined_mask.append(tuple(mask_row))
    supported, audit = v8.apply_pair_support_mask(baseline, tuple(combined_mask))
    return supported, active_plants, audit


def verify_fast_path(v9, v8, v6, v5, baseline: WeightedNetwork) -> None:
    for index, strength in enumerate((0.25, 0.5, 0.75)):
        support_seed = SEED + 90_000_000 + index
        weight_seed = SEED + 190_000_000 + index
        expected, _audit = v9.realize_local_context(
            baseline,
            support_seed=support_seed,
            support_strength=strength,
            weight_seed=weight_seed,
            weight_strength=0.5,
        )
        supported, _active, _fast_audit = fast_supported_context(
            v9, v8, v6, baseline,
            support_seed=support_seed,
            support_strength=strength,
        )
        actual = None if supported is None else v5.realize_local_context(
            supported,
            context_seed=weight_seed,
            context_strength=0.5,
        )
        if actual != expected:
            raise RuntimeError("optimized Martinique v9 path diverged from frozen v9 implementation")


def predictive_summary(design: dict, baseline: WeightedNetwork) -> dict:
    v9 = load_module(V9_SCRIPT, "martinique_v9_core")
    v8 = load_module(V8_SCRIPT, "martinique_v9_v8")
    v6 = load_module(V6_SCRIPT, "martinique_v9_v6")
    v5 = load_module(V5_SCRIPT, "martinique_v9_v5")
    verify_fast_path(v9, v8, v6, v5, baseline)

    p = design["v9_predictive_distribution"]
    support_strengths = [float(x) for x in p["support_strengths"]]
    weight_strengths = [float(x) for x in p["weight_strengths"]]
    replicates = int(p["replicates_per_support_x_weight_setting"])
    contexts_per_draw = int(p["contexts_per_predictive_draw"])
    baseline_pairs = pair_set(baseline)
    baseline_plants = set(baseline.plant_names)
    baseline_pollinators = set(baseline.pollinator_names)
    pooled_shannon = weighted_shannon(baseline)

    distributions = {
        "median_active_interacting_plant_fraction": [],
        "median_pair_support_fraction_given_active_plants": [],
        "interaction_shannon_relative_context_range": [],
    }
    secondary = {
        "mean_pairwise_jaccard_turnover": [],
        "active_pollinator_fraction_median": [],
        "zero_interaction_context_count": [],
    }
    setting_summary = {}

    for support_index, support_strength in enumerate(support_strengths):
        support_draws = []
        for replicate in range(replicates):
            supported_contexts = []
            active_plant_sets = []
            support_pair_sets = []
            pair_fractions = []
            plant_fractions = []
            pollinator_fractions = []
            for context_index in range(contexts_per_draw):
                support_seed = SEED + support_index * 10_000_000 + replicate * 10_000 + context_index
                supported, active_indices, _audit = fast_supported_context(
                    v9, v8, v6, baseline,
                    support_seed=support_seed,
                    support_strength=support_strength,
                )
                active_plants = {baseline.plant_names[index] for index in active_indices}
                pairs = pair_set(supported)
                denominator = {pair for pair in baseline_pairs if pair[0] in active_plants}
                plant_fractions.append(len(active_plants & baseline_plants) / len(baseline_plants))
                if denominator:
                    pair_fractions.append(len(pairs) / len(denominator))
                pollinators = {pollinator for _, pollinator in pairs}
                pollinator_fractions.append(len(pollinators) / len(baseline_pollinators))
                supported_contexts.append(supported)
                active_plant_sets.append(active_plants)
                support_pair_sets.append(pairs)
            if not pair_fractions:
                raise RuntimeError("synthetic draw has no defined conditional pair-support context")
            support_draws.append({
                "replicate": replicate,
                "supported_contexts": supported_contexts,
                "pair_sets": support_pair_sets,
                "plant_metric": float(statistics.median(plant_fractions)),
                "pair_metric": float(statistics.median(pair_fractions)),
                "turnover": jaccard_turnover(support_pair_sets),
                "pollinator_median": float(statistics.median(pollinator_fractions)),
                "empty_count": sum(not pairs for pairs in support_pair_sets),
            })

        for weight_index, weight_strength in enumerate(weight_strengths):
            setting_key = f"support={support_strength}|weight={weight_strength}"
            setting_values = {key: [] for key in distributions}
            for row in support_draws:
                shannons = []
                for context_index, supported in enumerate(row["supported_contexts"]):
                    if supported is None:
                        shannons.append(0.0)
                        continue
                    weight_seed = (
                        SEED + 500_000_000 + support_index * 20_000_000
                        + weight_index * 4_000_000 + row["replicate"] * 10_000 + context_index
                    )
                    realized = v5.realize_local_context(
                        supported,
                        context_seed=weight_seed,
                        context_strength=weight_strength,
                    )
                    shannons.append(weighted_shannon(realized))
                metrics = {
                    "median_active_interacting_plant_fraction": row["plant_metric"],
                    "median_pair_support_fraction_given_active_plants": row["pair_metric"],
                    "interaction_shannon_relative_context_range": (max(shannons) - min(shannons)) / pooled_shannon,
                }
                for key, value in metrics.items():
                    distributions[key].append(value)
                    setting_values[key].append(value)
                secondary["mean_pairwise_jaccard_turnover"].append(row["turnover"])
                secondary["active_pollinator_fraction_median"].append(row["pollinator_median"])
                secondary["zero_interaction_context_count"].append(float(row["empty_count"]))
            setting_summary[setting_key] = {key: envelope(values) for key, values in setting_values.items()}

    expected_draws = int(p["predictive_draw_count"])
    if any(len(values) != expected_draws for values in distributions.values()):
        raise RuntimeError("predictive draw count diverged from frozen design")
    return {
        "predictive_draw_count": expected_draws,
        "contexts_per_draw": contexts_per_draw,
        "equal_support_and_weight_setting_weights": True,
        "fast_path_identity_with_frozen_v9_verified": True,
        "primary_envelopes": {key: envelope(values) for key, values in distributions.items()},
        "secondary_envelopes": {key: envelope(values) for key, values in secondary.items()},
        "setting_summary": setting_summary,
    }


def run() -> dict:
    design = json.loads(DESIGN.read_text())
    reconstruction = json.loads(RECONSTRUCTION.read_text())
    if design["target_metrics_calculated"] is not False:
        raise RuntimeError("Martinique validation design is not pre-target frozen")
    empirical, baseline, _raw = build_empirical(design, reconstruction)
    structural = empirical["structural_gate"]
    if not structural["passes"]:
        return {
            "schema_version": "1.0",
            "analysis": "abm_v9_martinique_prospective_validation",
            "status": "blocked_before_target_predictive_adequacy",
            "target_metrics_inspected": False,
            "v9_predictive_distribution_calculated": False,
            "empirical_structure": empirical,
            "decision": design["pre_target_structural_gate"]["failure_decision"],
            "claim_boundary": "The frozen independent floral-opportunity measurement does not cover all observed interaction plant endpoints at the Site x Period scale, so full-v9 predictive adequacy is not evaluated and Martinique is not counted as a v9 biological falsification.",
        }
    assert baseline is not None
    predictive = predictive_summary(design, baseline)
    primary = empirical["primary_estimands"]
    tests = {
        key + "_adequacy": inside(value, predictive["primary_envelopes"][key])
        for key, value in primary.items()
    }
    survived = all(tests.values())
    decision = (
        "v9_survives_martinique_full_local_hierarchy_test"
        if survived else
        "v9_fails_martinique_full_local_hierarchy_predictive_adequacy"
    )
    return {
        "schema_version": "1.0",
        "analysis": "abm_v9_martinique_prospective_validation",
        "status": "held_out_full_v9_local_hierarchy_validation",
        "target_metrics_inspected": True,
        "v9_predictive_distribution_calculated": True,
        "empirical": empirical,
        "predictive": predictive,
        "tests": tests,
        "decision": decision,
        "claim_boundary": design["claim_boundary"],
    }


def main() -> None:
    payload = run()
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")
    summary = {
        "decision": payload["decision"],
        "status": payload["status"],
        "target_metrics_inspected": payload["target_metrics_inspected"],
    }
    if payload["target_metrics_inspected"]:
        summary["empirical_primary"] = payload["empirical"]["primary_estimands"]
        summary["predictive_primary_envelopes"] = payload["predictive"]["primary_envelopes"]
        summary["tests"] = payload["tests"]
    else:
        summary["structural_gate"] = payload["empirical_structure"]["structural_gate"]
    print(json.dumps(summary, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
